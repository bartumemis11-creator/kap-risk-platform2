# -*- coding: utf-8 -*-
"""
KAP Risk İzleme ve Erken Uyarı Platformu
=========================================
KAP (Kamuyu Aydınlatma Platformu, kap.org.tr) bildirimlerini tarayarak
şirketlerin kötüye gidişine işaret eden sinyalleri tespit eder, ağırlıklı
bir risk skoru üretir ve yapılandırılmış bir risk yönetimi raporu sunar.

İzlenen risk boyutları:
  • Temerrüt ve ödeme performansı bozulması
  • İflas, tasfiye, konkordato
  • Finansal yeniden yapılandırma, sermaye kaybı (TTK 376)
  • Yakın İzleme Pazarı'na alınma, kotasyon/işlem sırası riskleri
  • Kredi derecelendirme notunun kötüleşmesi (yön tespiti ile)
  • İcra ve haciz takipleri
  • Regülatör cezaları (SPK, BDDK, EPDK, Rekabet Kurumu, vergi)
  • İhaleye fesat, yolsuzluk, kayyum, adli süreçler
  • Denetçi görüşü ve işletme sürekliliği şüphesi
  • Önemli davalar ve tahkim
  • Piyasa tedbirleri (brüt takas, VBTS, açığa satış yasağı)
  • Faaliyet riski (üretim durdurma, grev, kaza)

Çalıştırma:   streamlit run kap_risk_app.py
Gereksinim:   streamlit, pandas, openpyxl  (pip install streamlit pandas openpyxl)
"""

import email.utils
import html as html_lib
import io
import json
import os
import random
import re
import threading
import time
import urllib.parse

import requests
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo

import altair as alt
import openpyxl
import pandas as pd
import streamlit as st

try:  # büyük taramalarda (tümünü seç) 5000 satır sınırına takılmamak için
    alt.data_transformers.disable_max_rows()
except Exception:
    pass
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ════════════════════════════════════════════════════════════ sabitler ════

BASE = "https://www.kap.org.tr"
TIMEOUT = 35
MAX_WORKERS = 4                # KAP'ı yormadan makul paralellik
MAX_SANE_RESULT = 800          # üstü: filtre uygulanmamış genel döküm
MIN_REQ_INTERVAL = 0.25        # istekler arası küresel asgari aralık (sn)
COOLDOWN_SECONDS = 8.0         # kısıtlama algılanınca küresel bekleme
EARLIEST_YEAR = 2015           # tarih seçicinin alt sınırı

STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          "kap_risk_izleme_gecmis.json")

try:
    ISTANBUL_TZ = ZoneInfo("Europe/Istanbul")
except Exception:
    ISTANBUL_TZ = timezone(timedelta(hours=3))

# Google News RSS sorgusunda aranan risk terimleri
NEWS_TERMS = ('(temerrüt OR konkordato OR iflas OR haciz OR "işlem yasağı" '
              'OR "yakın izleme" OR kayyum OR soruşturma OR "idari para '
              'cezası" OR "yeniden yapılandırma" OR tasfiye)')

TR_MAP = str.maketrans("çÇğĞıİöÖşŞüÜ", "ccggiioossuu")

SEVERITY_ORDER = {"KRİTİK": 0, "YÜKSEK": 1, "ORTA": 2, "DÜŞÜK": 3}
SEVERITY_EMOJI = {"KRİTİK": "🔴", "YÜKSEK": "🟠", "ORTA": "🟡", "DÜŞÜK": "🔵"}
GRADE_META = [
    (70, "E", "KRİTİK", "⚫", "#7f1d1d"),
    (45, "D", "YÜKSEK", "🔴", "#dc2626"),
    (20, "C", "ORTA", "🟠", "#ea580c"),
    (1,  "B", "DÜŞÜK", "🟡", "#ca8a04"),
    (0,  "A", "TEMİZ", "🟢", "#16a34a"),
]

# ═══════════════════════════════════════════════════ risk sınıflandırma ════
# Her kural: kategori kimliği, etiket, ağırlık (1-10), anahtar kelimeler.
# Kelimeler normalize edilmiş (küçük harf, TR karakterler sadeleştirilmiş)
# metinde aranır. Ağırlık 9-10 → KRİTİK, 7-8 → YÜKSEK, 5-6 → ORTA, ≤4 → DÜŞÜK.

RISK_CATEGORIES = {
    "iflas":      ("İflas / Tasfiye / Konkordato", 10),
    "temerrut":   ("Temerrüt / Ödeme Performansı", 10),
    "yakin_izleme": ("Yakın İzleme / Kotasyon Riski", 9),
    "ihale_yolsuzluk": ("İhaleye Fesat / Yolsuzluk / Adli", 9),
    "yapilandirma": ("Finansal Yeniden Yapılandırma", 8),
    "denetim":    ("Denetçi Görüşü / Süreklilik Şüphesi", 8),
    "derecelendirme": ("Kredi Notu Kötüleşmesi", 7),
    "icra":       ("İcra / Haciz Takipleri", 7),
    "regulator":  ("Regülatör Cezası / Yaptırım", 7),
    "faaliyet":   ("Faaliyet / Üretim Riski", 5),
    "yonetim":    ("Yönetim / Genel Kurul Riski", 5),
    "varlik_satisi": ("Varlık Satışı / Teminat-Rehin", 4),
    "karlilik":   ("Kârlılık / Temettü Sinyali", 4),
    "dava":       ("Önemli Dava / Tahkim", 4),
    "piyasa_tedbir": ("Piyasa Tedbirleri", 3),
}

RISK_RULES = [
    ("iflas", ["iflas", "tasfiye karari", "tasfiyeye gir", "tasfiye sureci",
               "konkordato", "borca batik"]),
    ("temerrut", ["temerrut", "temerrud", "odeme guclugu", "odenmemesi",
                  "odenememesi", "odeme yukumlulugunu yerine getirememe",
                  "vadesinde odenmeme", "kupon odemesi yapilamama",
                  "anapara odemesi yapilamama", "karsiliksiz cek", "protesto",
                  "odemelerini durdur", "borc odemede gecikme",
                  "odeyemez durum", "odenemez durum",
                  "odemesi gerceklestirileme", "odeme yapilamam",
                  "likidite sikisikligi", "nakit akisindaki bozulma",
                  "finansmana erisimde guclu", "odeme planinin revize",
                  # MKK/Takasbank borçlanma aracı ödeme aksaması bildirimleri
                  "gerceklesmeyen itfa", "gerceklesmeyen kupon",
                  "gerceklesmeyen getiri", "gerceklesmeyen odeme",
                  "itfa odemesinin gecik", "kupon odemesinin gecik",
                  "odemesinin ertelen", "erken itfa talebi",
                  "gayrinakdi finansman sikisik"]),
    ("yakin_izleme", ["yakin izleme", "islem sirasi kapatil",
                      "islem sirasinin kapatil", "islem sirasi durdur",
                      "kottan cikar", "borsa kotundan", "kotasyon sartlari",
                      "piyasa oncesi islem platformu", "poip",
                      # borçlanma araçları için gözaltı pazarı + VİOP'ta
                      # yeni vade ayı açılmaması gibi borsa güven sinyalleri
                      "gozalti pazari", "isleme acilmamas",
                      "isleme acilmayacak", "vade aylarinin acilmamas",
                      "sozlesmelerinde islemlerin durdurul"]),
    ("ihale_yolsuzluk", ["ihaleye fesat", "ihalelerden yasakla",
                         "ihale yasagi", "kamu ihalelerinden", "yolsuzluk",
                         "rusvet", "sahtecilik", "dolandiricilik",
                         "kayyum atan", "kayyum karari", "tutuklan",
                         "gozaltina alin", "el konul", "suc duyurusu"]),
    ("yapilandirma", ["finansal yeniden yapilandirma", "borclarin yeniden yapilandir",
                      "borc yapilandirma", "yeniden yapilandirma",
                      "sermaye kaybi", "ttk 376", "376 nci madde",
                      "376. madde", "teknik iflas", "borc erteleme",
                      "finansal zorluk", "operasyonel zorluk",
                      "finansal ve operasyonel",
                      # banka/faktoring kredilerinin yapılandırılması (MARTI),
                      # covenant/kredi şartı ihlali (SASA), zarar mahsuplu
                      # sermaye azaltımı
                      "kredi yapilandirma", "kredilerin yapilandiril",
                      "yapilandirma protokol", "kredilerinin raporlanmasi",
                      "covenant", "kredi sartlarinin ihlal",
                      "sermayenin azaltilmasi", "zarar mahsubu",
                      "zararlarin mahsubu"]),
    ("denetim", ["olumsuz gorus", "gorus bildirmekten kacin", "sartli gorus",
                 "isletmenin surekliligi", "surekliligine iliskin onemli belirsizlik",
                 "going concern"]),
    ("icra", ["icra takib", "icra takip", "icra dairesi", "icra mudurlugu",
              "haciz", "ihtiyati haciz", "aleyhine takip", "kanuni takip",
              "yasal takip baslat", "takip baslatil", "rehnin paraya cevril"]),
    ("regulator", ["idari para cezasi", "vergi cezasi", "vergi incelemesi",
                   "tarhiyat", "rekabet kurumu", "rekabet kurulu sorusturma",
                   "bddk", "epdk", "masak", "spk tarafindan ceza",
                   "kurul tarafindan verilen ceza", "para cezasi uygulan",
                   "yaptirim", "sorusturma acil", "sorusturma baslat",
                   "limit asimi", "sirketin uyaril", "uyari verilmesi"]),
    ("faaliyet", ["uretim durdur", "uretime ara", "faaliyet durdur",
                  "faaliyetlerin durdurulmasi", "fabrika kapat", "is durdurma",
                  "grev", "lokavt", "yangin", "patlama", "is kazasi",
                  "lisans iptal", "ruhsat iptal",
                  "sube kapanis", "subelerin kapat", "magaza kapanis"]),
    ("yonetim", ["istifa", "gorevinden ayril", "gorevden alin",
                 "genel mudur degisik", "yonetim kurulu uyeliginden",
                 "nisabinin saglanamamas", "nisap saglanamadigi",
                 "genel kurul toplantisinin ertelen", "genel kurulun ertelen",
                 "genel kurulu ertelen", "genel kurul tehiri",
                 "genel kurulun tehir",
                 # üst kademe ayrılıkları/değişimleri ve hakim ortak satışı
                 "uyesi ayrilmasi", "uyesinin ayrilmasi",
                 "yardimcisinin ayrilmasi", "genel mudurun ayrilmasi",
                 "baskaninin ayrilmasi", "ust yonetici degisikli",
                 "ust yonetim degisikli", "yonetim kurulu degisikli",
                 "yonetim kurulu baskani ve uye degisikli",
                 "ortaklar pay satisi", "hakim ortagin pay satis"]),
    ("karlilik", ["kar dagitilmamasi", "kar payi dagitilmamasi",
                  "kar dagitimi yapilmamasi", "temettu odenmemesi",
                  "donem zarari nedeniyle", "zarar edilmesi nedeniyle"]),
    ("varlik_satisi", ["duran varlik satis", "gayrimenkul satisi",
                       "tasinmaz satisi", "istirak satisi", "istirak paylarinin satisi",
                       "bagli ortaklik satisi", "bagli ortaklik paylarinin satisi",
                       "varliklarin satisi", "maddi duran varlik satis",
                       "arsa satisi", "fabrika satisi",
                       "tahsili gecikmis alacak",
                       # "satım" yazım varyantı (piyasa taramasında kaçtı:
                       # 'Maddi Duran Varlık Satımı — ofis satışı')
                       "duran varlik satim", "varlik satimi",
                       "gayrimenkul satimi", "tasinmaz satimi",
                       # borç teminatı olarak ipotek/rehin tesisi
                       "ipotek tesis", "rehin tesis", "ipotek verilmesi",
                       "rehin verilmesi", "ipotek edilmesi"]),
    ("dava", ["dava acil", "dava acti", "aleyhine dava", "aleyhine acilan dava",
              "davanin kabul", "tahkim", "arabuluculuk basvuru",
              "tazminat davasi"]),
    ("piyasa_tedbir", ["brut takas", "volatilite bazli tedbir", "vbts",
                       "aciga satis yasagi", "kredili islem yasagi",
                       "tek fiyat yontemi", "islem yasagi", "tedbir karari",
                       "tedbir uygulan", "yatirim araci bazinda tedbir",
                       "pazar degisikligi", "olagan disi fiyat"]),
]

# Derecelendirme bildirimleri ayrı ele alınır (yön tespiti gerekir)
RATING_TRIGGERS = ["derecelendirme", "kredi notu", "rating", "gorunum"]
RATING_NEGATIVE = ["dusurul", "dusurdu", "indirdi", "indirilm", "indirim",
                   "negatif", "asagi yonlu", "geri cek", "izlemeye al",
                   "not cekil", "durdurul", "askiya al", "temerrut",
                   "kotules", "d seviyesi", "default"]
RATING_POSITIVE = ["yukselt", "teyit", "korundu", "duragan", "stabil",
                   "pozitif", "yukari yonlu",
                   # derecelendirme SÖZLEŞMESİ imzalanması nötr/olumludur
                   "sozlesme imzalan", "sozlesmesi imzalan",
                   "sozlesmesinin imzalan", "anlasma imzalan"]

# Olumlu yön: KATEGORİYE ÖZGÜ iyileşme kalıpları. Genel bir listeyle
# arama yapmak riskliydi — "konkordato mühleti sona erdi, iflas başladı"
# gibi ağır bir bildirim, metindeki "sona erdi" yüzünden iyileşme sayılıp
# skor dışı kalıyordu. Kalıp yalnızca eşleşen kategorinin bağlamında
# aranır. İflas/yapılandırma gibi kategorilerde kısayol hiç yoktur;
# temerrütte ise yalnızca borcun FİİLEN ödendiğini bildiren dar kalıplar
# kabul edilir (ödeme gerçekleşmişse temerrüt gerçekten çözülmüştür).
IMPROVEMENT_HINTS_BY_CAT = {
    "yakin_izleme": ["yakin izleme pazarindan cikar", "ana pazara gec",
                     "ana pazara alin", "gozalti pazarindan cikar",
                     "yeniden islem gorme"],
    "piyasa_tedbir": ["tedbirin kaldiril", "yasagin kaldiril",
                      "kaldirilmasina karar", "tedbir sona erdi"],
    "regulator": ["sorusturmanin sonlandiril", "sorusturmasinin sonlandiril",
                  "sorusturmanin kapatil", "sorusturmasinin kapatil",
                  "sorusturmanin sona", "sorusturmasinin sona",
                  "ceza verilmemesi", "sorusturma acilmamasina"],
    "dava": ["davanin lehine sonuclan", "davanin reddi", "lehte sonuclan",
             "davadan feragat"],
    "icra": ["takipten vazgec", "takibin iptal", "haczin kaldiril",
             "borcun odendi", "tamamen odendi"],
    "temerrut": ["borcun odendi", "tamamen odendi",
                 "odemenin gerceklestirildigi"],
}

# Gürültü: risk değeri taşımayan rutin bildirimler.
# DİKKAT: "kar payı dağıtım" ve "genel kurul işlemlerine" BİLEREK yok —
# nisab sağlanamayan/ertelenen genel kurullar bu başlıklar altında gelir.
NOISE_PATTERNS = ["devre kesici", "endeks sirketlerinde degisiklik",
                  "endeks degisikligi",
                  "kurumsal yonetim uyum", "surdurulebilirlik raporu",
                  # yasaklı YATIRIMCILARIN paylarına ilişkin rutin MKK
                  # operasyon duyurusu — şirket riski değildir
                  "islem yasagi nedeniyle pay duyurusu",
                  # planlı/periyodik bakım duruşu üretim riski sayılmaz
                  "planli bakim", "periyodik bakim"]

# BIST/SPK kaynaklı piyasa geneli duyurular (şirketin kendi beyanı değil)
MARKET_WIDE_PUBLISHERS = ["BORSA İSTANBUL", "SERMAYE PİYASASI KURULU",
                          "MERKEZİ KAYIT", "TAKASBANK"]

RECENCY_STEPS = [(90, 1.0), (365, 0.75), (730, 0.5)]
RECENCY_FLOOR = 0.3
SCORE_GAIN = 2.4  # bulgu ağırlığı → skor katkısı çarpanı

# ═══════════════════════════════════════════════════════════ yardımcılar ════


def norm(s: str) -> str:
    if not s:
        return ""
    s = re.sub(r"['’`´]", "", s)   # kesme işaretleri eşleşmeyi bozmasın
    return s.translate(TR_MAP).lower()


# ── küresel hız sınırlayıcı + devre kesici ─────────────────────────────
# KAP art arda gelen istekleri geçici olarak kısıtlar; en sağlam çözüm
# istek yağmurunu hiç başlatmamaktır: tüm thread'ler tek bir zamanlayıcı
# üzerinden sıraya girer, kısıtlama algılanınca herkes birlikte bekler.
_rate_lock = threading.Lock()
_next_slot = [0.0]
_cooldown_until = [0.0]


def _throttle():
    while True:
        with _rate_lock:
            now = time.time()
            wait_cd = _cooldown_until[0] - now
            if wait_cd <= 0:
                wait = max(0.0, _next_slot[0] - now)
                _next_slot[0] = max(now, _next_slot[0]) + MIN_REQ_INTERVAL
        if wait_cd > 0:
            time.sleep(min(wait_cd, 1.0))
            continue
        if wait > 0:
            time.sleep(wait)
            with _rate_lock:   # slot beklerken cooldown tetiklenmiş olabilir
                if _cooldown_until[0] > time.time():
                    continue
        return


def _trigger_cooldown():
    with _rate_lock:
        _cooldown_until[0] = max(_cooldown_until[0],
                                 time.time() + COOLDOWN_SECONDS)


# keep-alive bağlantı havuzu: her istekte TLS el sıkışması tekrarlanmaz
_session = requests.Session()
_session.headers.update({
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                   "AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/126.0.0.0 Safari/537.36"),
    "Accept": "*/*", "Accept-Language": "tr",
})


def http_get(url: str, retries: int = 4) -> str:
    """KAP istekleri: küresel hız sınırlayıcı + devre kesiciden geçer."""
    last = None
    for attempt in range(retries):
        _throttle()
        try:
            resp = _session.get(url, timeout=TIMEOUT)
            resp.raise_for_status()
            resp.encoding = "utf-8"
            return resp.text
        except Exception as exc:
            last = exc
            # üstel geri çekilme + rastgele sapma (thread'ler senkron
            # yeniden denemesin diye); son denemeden sonra beklemek anlamsız
            if attempt < retries - 1:
                time.sleep(1.2 * (attempt + 1) + random.uniform(0, 0.8))
    raise RuntimeError(f"KAP isteği başarısız: {url} → {last}")


def plain_get(url: str, retries: int = 2) -> str:
    """KAP dışı kaynaklar (Google News): KAP'ın hız sınırlayıcısına ve
    devre kesicisine takılmadan, kendi hafif yeniden denemesiyle çeker."""
    last = None
    for attempt in range(retries):
        try:
            resp = _session.get(url, timeout=TIMEOUT)
            resp.raise_for_status()
            resp.encoding = "utf-8"
            return resp.text
        except Exception as exc:
            last = exc
            if attempt < retries - 1:
                time.sleep(1.0 + random.uniform(0, 0.5))
    raise RuntimeError(f"İstek başarısız: {url} → {last}")


def extract_flight_array(page: str, anchor: str = '\\"data\\":['):
    """Next.js SSR sayfasına gömülü (escape edilmiş) JSON dizisini çıkarır.

    None → sayfada veri bloğu hiç yok (bozuk/kısıtlanmış yanıt);
    []   → veri bloğu var ama boş (şirketin o dönemde bildirimi yok).
    """
    start = page.find(anchor)
    if start < 0:
        return None
    region = page[start:start + 3_000_000]
    txt = (region.replace('\\\\"', "@@Q@@")
                 .replace('\\"', '"')
                 .replace("@@Q@@", '\\"'))
    a = txt.find("[")
    depth, in_str, esc = 0, False, False
    for i, ch in enumerate(txt[a:], a):
        if in_str:
            if esc:
                esc = False
            elif ch == "\\":
                esc = True
            elif ch == '"':
                in_str = False
            continue
        if ch == '"':
            in_str = True
        elif ch == "[":
            depth += 1
        elif ch == "]":
            depth -= 1
            if depth == 0:
                try:
                    return json.loads(txt[a:i + 1])
                except json.JSONDecodeError:
                    # ayrıştırılamayan veri bloğu "temiz şirket" DEĞİL,
                    # bozuk yanıttır → None ile yeniden deneme tetiklenir
                    return None
    return None                       # dizi hiç kapanmadı: kesik yanıt


def parse_date(s: str) -> datetime:
    try:
        return datetime.strptime(s, "%d.%m.%Y %H:%M:%S")
    except Exception:
        try:
            return datetime.strptime(s[:10], "%d.%m.%Y")
        except Exception:
            return datetime.min


def recency_factor(dt: datetime) -> float:
    if dt == datetime.min:
        return RECENCY_FLOOR
    age = (datetime.now() - dt).days
    for limit, factor in RECENCY_STEPS:
        if age <= limit:
            return factor
    return RECENCY_FLOOR


def severity_for_weight(w: int) -> str:
    if w >= 9:
        return "KRİTİK"
    if w >= 7:
        return "YÜKSEK"
    if w >= 5:
        return "ORTA"
    return "DÜŞÜK"


def grade_for_score(score: float):
    for threshold, grade, label, emoji, color in GRADE_META:
        if score >= threshold:
            return grade, label, emoji, color
    return GRADE_META[-1][1:]

# ═══════════════════════════════════════════════════════════ veri katmanı ════


@st.cache_data(ttl=24 * 3600, show_spinner=False)
def fetch_member_directory() -> pd.DataFrame:
    """KAP üye rehberini çeker → hisse kodu/etiket, unvan, oid, işlem durumu.

    Toleranslı ayrıştırma: sayfadaki kaçışlı ve kaçışsız JSON kümelerinin
    ikisini de okur; hisse kodu olmayan ihraççılar ve payları işlem
    görmeyen (payIslemDurumu=0) üyeler de dahil edilir — temerrüt takibi
    için en kritik kesim genellikle bunlardır.
    """
    page = http_get(f"{BASE}/tr/bildirim-sorgu")
    txt = page.replace('\\"', '"')
    rows, seen_oid, seen_label = [], set(), set()
    for chunk in txt.split('{"kapMemberOid"')[1:]:
        chunk = chunk[:2500]
        oid = re.search(r'"mkkMemberOid":"([0-9a-fA-F]+)"', chunk)
        ttl = re.search(r'"kapMemberTitle":"(.*?)"', chunk)
        if not (oid and ttl) or oid.group(1) in seen_oid:
            continue
        seen_oid.add(oid.group(1))
        title = ttl.group(1)
        try:
            title = json.loads('"' + title.replace('"', '\\"') + '"')
        except json.JSONDecodeError:
            pass
        stk = re.search(r'"stockCode":"(.*?)"', chunk)
        pis = re.search(r'"payIslemDurumu":"(.*?)"', chunk)
        codes = [c.strip().upper() for c in
                 (stk.group(1) if stk else "").replace(";", ",").split(",")
                 if c.strip() and c.strip() != "-"]
        if codes:
            # birincil kod: en uzunu (YKB,YKBNK → YKBNK)
            code = max(codes, key=len)
            kodlar = ",".join(codes)
        else:
            # kodsuz ihraççı: unvandan kısa etiket üret (* işaretli)
            code = "*" + norm(title).split()[0].upper()[:9]
            kodlar = code
        label = code
        n = 2
        while label in seen_label:          # etiket çakışmasını gider
            label = f"{code}{n}"
            n += 1
        seen_label.add(label)
        rows.append({"hisse": label, "kodlar": kodlar, "unvan": title,
                     "oid": oid.group(1),
                     "islem": pis.group(1) if pis else ""})
    df = pd.DataFrame(rows)
    return df.sort_values("hisse").reset_index(drop=True)


# Yıl bazlı önbellek: KAPANMIŞ yılların bildirimleri değişmez → süresiz
# saklanır; güncel yıl ve güncel pencere 1 saatte bir tazelenir. Böylece
# saat başı otomatik yenileme yalnızca güncel dönemi çeker.
_year_cache: dict = {}
_year_lock = threading.Lock()
_YEAR_TTL_OPEN = 3600
_YEAR_CACHE_MAX = 12_000


def _year_cache_get(oid: str, yr):
    with _year_lock:
        hit = _year_cache.get((oid, yr))
    if not hit:
        return None
    ts, items = hit
    closed = yr is not None and yr < datetime.now().year
    if closed or time.time() - ts < _YEAR_TTL_OPEN:
        return items
    return None


def _year_cache_put(oid: str, yr, items):
    with _year_lock:
        if len(_year_cache) > _YEAR_CACHE_MAX:   # bellek emniyeti
            _year_cache.clear()
        _year_cache[(oid, yr)] = (time.time(), items)


def _fetch_year(oid: str, yr):
    """Tek yıl sorgusu → (bildirim listesi, başarı bayrağı).

    KAP geçici hız sınırlaması uyguladığında sayfa 200 döner ama veri
    bloğu bulunmaz; bu durum başarısızlık sayılır ve yeniden denenir.
    """
    cached = _year_cache_get(oid, yr)
    if cached is not None:
        return cached, True
    url = (f"{BASE}/tr/bildirim-sorgu-sonuc?srcbar=Y&cmp=Y&cat=2"
           f"&m={oid}&t=X&slf=ALL")
    if yr:
        url += f"&yr={yr}"
    for attempt in range(3):
        try:
            arr = extract_flight_array(http_get(url))
        except RuntimeError:
            arr = None
        if arr is not None:
            if len(arr) > MAX_SANE_RESULT:
                # filtre çalışmamış genel döküm: veriyi kullanma ama bunu
                # "başarı" da sayma — kullanıcı 'kısmi veri' uyarısı görür
                return [], False
            _year_cache_put(oid, yr, arr)
            return arr, True
        _trigger_cooldown()                  # kısıtlama: herkes beklesin
        time.sleep(2.5 * (attempt + 1))
    return [], False


def fetch_company_disclosures(oid: str, years: tuple):
    """Şirket bildirimlerini seçilen yıllar + güncel pencere için çeker.

    → (bildirim listesi, başarısız sorgu sayısı). Önbellek yıl bazındadır
    ve yalnızca başarılı çekimler saklanır; başarısız sorgular bir sonraki
    taramada otomatik yeniden denenir.
    """
    queries = list(years) + [None]
    seen, fails = {}, 0
    with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, len(queries))) as ex:
        futures = [ex.submit(_fetch_year, oid, yr) for yr in queries]
        for fut in as_completed(futures):
            items, ok = fut.result()
            if not ok:
                fails += 1
            for it in items:
                basic = it.get("disclosureBasic") or it
                idx = basic.get("disclosureIndex")
                if idx and idx not in seen:
                    seen[idx] = basic
    result = sorted(seen.values(),
                    key=lambda b: parse_date(b.get("publishDate", "")),
                    reverse=True)
    return result, fails


_detail_cache: dict = {}
_detail_lock = threading.Lock()


def fetch_detail_text(disclosure_index: int):
    """Bildirim detay sayfasından tam açıklama metnini çıkarır.

    None → çekim başarısız (kısıtlama/ağ); "" → sayfa geldi ama metin yok.
    Worker thread'lerden çağrıldığı için st.cache_data yerine kilitli
    modül önbelleği kullanır (Streamlit ScriptRunContext gerektirmez).
    """
    with _detail_lock:
        if disclosure_index in _detail_cache:
            return _detail_cache[disclosure_index]
    try:
        page = http_get(f"{BASE}/tr/Bildirim/{disclosure_index}")
    except RuntimeError:
        return None                      # başarısız → önbelleklenmez
    txt = re.sub(r"\\u([0-9a-fA-F]{4})",
                 lambda m: chr(int(m.group(1), 16)), page)
    start = txt.find("taxonomy")
    if start < 0:
        start = txt.find("summaryInfo")
    if start < 0:
        return ""
    end = txt.find("footerNote", start)
    if end < 0:
        end = min(len(txt), start + 200_000)
    seg = txt[max(0, start - 300):end]
    seg = re.sub(r"<[^>]+>", " ", seg)
    seg = html_lib.unescape(seg)
    seg = re.sub(r"\\[ntr]", " ", seg)
    seg = re.sub(r"[\\\"{}\[\]]", " ", seg)
    seg = re.sub(r"oda_[A-Za-z]+\|?", " ", seg)   # taksonomi alan adları
    seg = re.sub(r"\s+", " ", seg)
    result = seg[:25_000]
    with _detail_lock:
        _detail_cache[disclosure_index] = result
        if len(_detail_cache) > 5000:   # bellek emniyeti
            _detail_cache.clear()
    return result

# ═══════════════════════════════════════════════════════════ risk motoru ════


def classify(basic: dict, member: dict, detail_text: str = "") -> dict | None:
    """Tek bildirimi sınıflandır. None → risk sinyali değil."""
    surface = " | ".join(filter(None, [
        basic.get("title"), basic.get("summary"), basic.get("subject"),
        basic.get("ruleTypeTerm")]))
    text = norm(surface)
    deep = norm(detail_text) if detail_text else ""
    combined = text + " ~ " + deep
    if not text.strip():
        return None
    for noise in NOISE_PATTERNS:
        if noise in text:
            return None

    publisher = norm(basic.get("companyTitle") or "")
    is_market_wide = any(norm(p) in publisher for p in MARKET_WIDE_PUBLISHERS) \
        and norm(member["unvan"])[:15] not in publisher

    # 0) Türetilmiş sinyal: finansal raporun dönem sonundan çok geç
    #    yayımlanması (metinde anahtar kelime yoktur; tarih aritmetiği).
    #    Yıl sonu raporu için ~130 gün, ara dönem için ~75 gün eşik.
    title_n = norm(basic.get("title") or "")
    if title_n.startswith(("finansal rapor", "sorumluluk beyani",
                           "faaliyet raporu")):
        pub = parse_date(basic.get("publishDate", ""))
        ends = []
        for dm, yy in re.findall(r"(31\.03|30\.06|30\.09|31\.12)\.(\d{4})",
                                 surface):
            try:
                ends.append(datetime.strptime(f"{dm}.{yy}", "%d.%m.%Y"))
            except ValueError:
                pass
        past = [e for e in ends if e < pub]
        if pub != datetime.min and past:
            pe = max(past)
            delay = (pub - pe).days
            # Yıllık rapor tespiti: 31.12 kapanışı, "özel hesap dönemi"
            # ibaresi veya ~12 aylık tarih aralığı (örn. MARTI'nın
            # 01.04–31.03 özel hesap dönemi) → yıllık eşik uygulanır.
            annual = pe.month == 12 or "ozel hesap donemi" in text
            if not annual:
                for s_d, e_d in re.findall(
                        r"(\d{2}\.\d{2}\.\d{4})\s*[-–]\s*(\d{2}\.\d{2}\.\d{4})",
                        surface):
                    try:
                        d1_ = datetime.strptime(s_d, "%d.%m.%Y")
                        d2_ = datetime.strptime(e_d, "%d.%m.%Y")
                    except ValueError:
                        continue
                    if d2_ == pe and (d2_ - d1_).days >= 300:
                        annual = True
                        break
            limit = 130 if annual else 75
            if delay > limit:
                return _finding(
                    basic, member, "denetim", "ORTA", 6,
                    f"finansal raporlama gecikmesi: dönem sonundan {delay} "
                    f"gün sonra yayımlandı (olağan eşik ~{limit} gün) — "
                    "denetim/kapanış sorunlarına işaret edebilir",
                    improvement=False, market_wide=is_market_wide)

    # 1) Derecelendirme: yön analizi — başlık/özet VE detay metni birlikte
    #    aranır (yalnız detaya bakmak, başlıktaki "notun düşürülmesi"
    #    ifadesini kaçırıyordu)
    if any(k in text for k in RATING_TRIGGERS):
        probe = combined
        neg = [k for k in RATING_NEGATIVE if k in probe]
        pos = [k for k in RATING_POSITIVE if k in probe]
        if neg and (not pos or len(neg) >= len(pos)):
            return _finding(basic, member, "derecelendirme", "KRİTİK", 9,
                            f"olumsuz not aksiyonu tespit edildi ({', '.join(neg[:3])})",
                            improvement=False, market_wide=is_market_wide)
        if pos and not neg:
            return None                      # teyit / yükseltme → risk değil
        if deep:
            return None                      # detay okundu, olumsuzluk yok
        return _finding(basic, member, "derecelendirme", "DÜŞÜK", 3,
                        "derecelendirme bildirimi — yön özetten belirlenemedi "
                        "(rutin yıllık not yayını olabilir), derin analiz veya "
                        "detay teyidi önerilir", improvement=False,
                        market_wide=is_market_wide)

    # 2) Genel kurallar (önce yüzey metni, derin modda tam metin)
    for cat_id, keywords in RISK_RULES:
        label, weight = RISK_CATEGORIES[cat_id]
        for kw in keywords:
            hit_surface = kw in text
            hit_deep = deep and kw in deep
            if not (hit_surface or hit_deep):
                continue
            # iyileşme yalnızca EŞLEŞEN kategorinin kendi kalıplarıyla
            # tespit edilir; iflas/yapılandırma gibi ağır kategorilerde
            # kısayol yoktur (bkz. IMPROVEMENT_HINTS_BY_CAT açıklaması)
            cat_hints = IMPROVEMENT_HINTS_BY_CAT.get(cat_id, ())
            if any(h in combined for h in cat_hints):
                return _finding(basic, member, cat_id, "İYİLEŞME", 0,
                                f"olumlu yönlü gelişme ('{kw}' bağlamında "
                                "kaldırma/çıkarma/lehte sonuç ifadesi)",
                                improvement=True, market_wide=is_market_wide)
            w = weight if hit_surface else max(1, weight - 1)
            if is_market_wide and cat_id == "piyasa_tedbir":
                w = max(1, w - 1)
            sev = severity_for_weight(w)
            src = "başlık/özet" if hit_surface else "bildirim tam metni"
            note = f"'{kw}' ifadesi ({src})"
            if is_market_wide:
                note += " — BIST/SPK piyasa duyurusu"
            return _finding(basic, member, cat_id, sev, w, note,
                            improvement=False, market_wide=is_market_wide)
    return None


def _finding(basic, member, cat_id, severity, weight, reason,
             improvement, market_wide):
    label, _ = RISK_CATEGORIES[cat_id]
    dt = parse_date(basic.get("publishDate", ""))
    return {
        "hisse": member["hisse"],
        "sirket": member["unvan"],
        "oid": member.get("oid", ""),   # 🆕 anahtarı için kararlı kimlik
        "tarih": dt,
        "tarih_str": basic.get("publishDate", "")[:16],
        "kategori_id": cat_id,
        "kategori": label,
        "siddet": severity,
        "agirlik": weight,
        "guncellik": recency_factor(dt),
        "baslik": basic.get("title") or "",
        "ozet": (basic.get("summary") or "")[:500],
        "gerekce": reason,
        "iyilesme": improvement,
        "piyasa_geneli": market_wide,
        "bildirim_no": basic.get("disclosureIndex"),
        "link": f"{BASE}/tr/Bildirim/{basic.get('disclosureIndex')}",
    }


def _needs_detail(basic: dict) -> bool:
    """Derin modda hangi bildirimlerin tam metni çekilsin?"""
    text = norm(" | ".join(filter(None, [
        basic.get("title"), basic.get("summary"), basic.get("subject")])))
    if any(n in text for n in NOISE_PATTERNS):
        return False
    if any(k in text for k in RATING_TRIGGERS):
        return True
    # "Faaliyetlerin Kısmen veya Tamamen Durdurulması" şablonu başlıkta
    # "açıklama" içermediğinden derin okumaya girmiyordu — "durdurul" eklendi
    broad = ["ozel durum aciklamasi", "duyuru", "aciklama", "durdurul"]
    all_kw = [kw for _, kws in RISK_RULES for kw in kws]
    return any(k in text for k in all_kw) or any(b in text for b in broad)


# Varsayılan izleme evreni: BIST örnekleri + kullanıcının grup listesi.
# Gruplar unvan eşleşmesiyle çözülür; KAP üyesi olmayanlar yalnızca
# medya (Google News) tarafında izlenir.
DEFAULT_TICKERS = ["KONTR", "MARTI", "GSRAY", "FENER", "VESTL", "SASA",
                   "IHLAS", "THYAO", "TUPRS", "YKBNK"]
DEFAULT_GROUPS = ["LİMAK", "YAPI", "EKOPARK", "ROKETSAN", "BMC", "TOSYALI",
                  "TAŞYAPI", "CCN", "CEVAHİR", "ALBAYRAK", "VESTEL",
                  "BG MADEN", "ARCA", "ZORLU", "DOĞUŞ", "PASİFİK",
                  "MEDİPOLİTAN", "FINE", "KASTAMONU", "TOSÇELİK", "YENİKÖY",
                  "YILDIZLAR", "TÜRKERLER", "MELİKE TEKSTİL", "YATAĞAN",
                  "SÖĞÜT", "EREĞLİ", "GALERİ", "SAFİ", "CABA", "CENGİZ",
                  "KALYON", "TVF", "AKSA", "RAM", "KOLİN", "ÇELİKLER"]
GENERIC_GROUP_TERMS = {"yapi"}   # tek kelime olarak fazla genel


def resolve_default_members(directory: pd.DataFrame):
    """Varsayılan evreni çöz → (oid listesi, KAP'ta bulunamayan terimler).

    Tek kelimelik terimler unvanın ilk 4 kelimesinden biriyle birebir
    eşleşmeli (DOĞUŞ ↛ DOĞUSAN); çok kelimeliler alt dizge olarak aranır.
    Grup adını taşıyan bağımsız denetim/YMM firmaları elenir.
    """
    oids, unmatched = [], []
    tick = set(DEFAULT_TICKERS)
    for row in directory.itertuples():
        if tick & set(str(row.kodlar).split(",")):
            oids.append(row.oid)
    for term in DEFAULT_GROUPS:
        tn = norm(term)
        hit = False
        for row in directory.itertuples():
            un = norm(row.unvan)
            if "bagimsiz denetim" in un or "yeminli mali" in un:
                continue
            if " " in tn:
                ok = tn in un
            elif tn in GENERIC_GROUP_TERMS:
                # fazla genel terim: yalnızca aktif üyelerde ve unvan
                # başında ara (eski/kapalı iştirak enflasyonunu önler)
                ok = un.startswith(tn + " ") and str(row.islem) == "1"
            else:
                ok = tn in re.split(r"[\s.,\-]+", un)[:4]
            if ok:
                oids.append(row.oid)
                hit = True
        if not hit:
            unmatched.append(term)
    seen, out = set(), []
    for o in oids:
        if o not in seen:
            seen.add(o)
            out.append(o)
    return out, unmatched


def company_short_name(unvan: str) -> str:
    """Haber aramasında kullanılacak kısa şirket adı (ör. 'SASA POLYESTER')."""
    words, out = unvan.split(), []
    for w in words:
        # 'A.Ş.' → 'as', 'T.A.Ş.' → 'tas', 'A.O.' → 'ao' (tüm noktalar atılır)
        if norm(w).replace(".", "") in ("as", "ao", "tas", "tao", "ve"):
            if len("".join(out)) >= 8:
                break
        out.append(w)
        if len("".join(out)) >= 10 or len(out) >= 3:
            break
    return " ".join(out) if out else unvan.split()[0]


_news_cache: dict = {}
_news_lock = threading.Lock()


def fetch_company_news(short: str, max_items: int = 12) -> list:
    """Google News RSS'ten şirket için risk temalı haberleri çeker."""
    with _news_lock:
        hit = _news_cache.get(short)
        if hit and time.time() - hit[0] < 3600:
            return hit[1]
    q = urllib.parse.quote(f'"{short}" {NEWS_TERMS}')
    url = f"https://news.google.com/rss/search?q={q}&hl=tr&gl=TR&ceid=TR:tr"
    try:
        # KAP dışı kaynak: KAP hız sınırlayıcısına takılmadan çek
        xml_text = plain_get(url, retries=2)
        root = ET.fromstring(xml_text)
    except (RuntimeError, ET.ParseError):
        return []
    items = []
    for it in root.iter("item"):
        title = (it.findtext("title") or "").strip()
        link = (it.findtext("link") or "").strip()
        src = (it.findtext("source") or "").strip()
        pub = (it.findtext("pubDate") or "").strip()
        try:
            dt = (email.utils.parsedate_to_datetime(pub)
                  .astimezone(timezone(timedelta(hours=3)))   # TR saati
                  .replace(tzinfo=None))
        except Exception:
            dt = datetime.min
        if title and link:
            items.append({"baslik": title, "link": link, "kaynak": src,
                          "dt": dt,
                          "tarih_str": dt.strftime("%d.%m.%Y")
                          if dt != datetime.min else ""})
    items.sort(key=lambda x: x["dt"], reverse=True)
    items = items[:max_items]
    with _news_lock:
        _news_cache[short] = (time.time(), items)
    return items


def scan_company(member: dict, years: tuple, deep: bool,
                 date_range: tuple = None) -> dict:
    """Bir şirketi tarar → bulgular + istatistik.

    date_range=(başlangıç, bitiş) verilirse bildirimler bu aralığa
    (yayın tarihine göre) daraltılır — KAP tarih parametresini sunucu
    tarafında desteklemediği için filtre yerelde uygulanır.
    """
    disclosures, fetch_fails = fetch_company_disclosures(member["oid"], years)
    if date_range:
        d0, d1 = date_range
        lo = datetime.combine(d0, datetime.min.time())
        hi = datetime.combine(d1, datetime.max.time())
        kept = []
        for b in disclosures:
            dt = parse_date(b.get("publishDate", ""))
            # tarihi çözülemeyen kayıtları sessizce eleme — taramada kalsın
            if dt == datetime.min or lo <= dt <= hi:
                kept.append(b)
        disclosures = kept
    details, detail_fails = {}, 0
    if deep:
        cands = [b for b in disclosures if _needs_detail(b)][:60]
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            futs = {ex.submit(fetch_detail_text, b["disclosureIndex"]): b["disclosureIndex"]
                    for b in cands if b.get("disclosureIndex")}
            for fut in as_completed(futs):
                txt = fut.result()
                if txt is None:          # kısıtlama → derin analiz eksik
                    detail_fails += 1
                    txt = ""
                details[futs[fut]] = txt
    findings, improvements = [], []
    for basic in disclosures:
        res = classify(basic, member,
                       details.get(basic.get("disclosureIndex"), ""))
        if res is None:
            continue
        (improvements if res["iyilesme"] else findings).append(res)

    # Skor: aynı kategorideki tekrar bulgular azalan ağırlıkla sayılır
    # (tek olay birden çok bildirim doğurur; şişmeyi önler)
    score = 0.0
    by_cat: dict = {}
    for f in findings:
        by_cat.setdefault(f["kategori_id"], []).append(f)
    for cat_findings in by_cat.values():
        cat_findings.sort(key=lambda f: -(f["agirlik"] * f["guncellik"]))
        for rank, f in enumerate(cat_findings):
            damp = 1.0 if rank == 0 else 0.35
            score += f["agirlik"] * f["guncellik"] * SCORE_GAIN * damp
    score = min(100.0, score)
    grade, glabel, gemoji, gcolor = grade_for_score(score)
    if findings and grade == "A":
        # bulgusu olan şirket asla "TEMİZ" etiketi almasın
        grade, glabel, gemoji, gcolor = "B", "DÜŞÜK", "🟡", "#ca8a04"
    result = {
        "member": member, "taranan": len(disclosures),
        "findings": findings, "improvements": improvements,
        "skor": round(score, 1), "not": grade, "seviye": glabel,
        "emoji": gemoji, "renk": gcolor,
        "veri_hatasi": fetch_fails + detail_fails,
    }
    if not disclosures and fetch_fails:
        # hiç veri alınamadı → "temiz" değil, veri sorunu
        result.update({"not": "-", "seviye": "VERİ EKSİK / KONTROL GEREKLİ",
                       "emoji": "⚠️", "renk": "#64748b"})
    return result

# ═══════════════════════════════════════════════════════════ excel raporu ════

try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except ImportError:                       # openpyxl iç API değişirse
    ILLEGAL_CHARACTERS_RE = re.compile(
        r"[\000-\010]|[\013-\014]|[\016-\037]")


def _xl(v):
    """Excel hücresine yazılamayan kontrol karakterlerini ayıkla."""
    if isinstance(v, str):
        return ILLEGAL_CHARACTERS_RE.sub("", v)
    return v


def _xlrow(vals):
    return [_xl(v) for v in vals]


THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F3864")
SEV_FILL = {
    "KRİTİK": PatternFill("solid", fgColor="F8CBCB"),
    "YÜKSEK": PatternFill("solid", fgColor="FCE0C8"),
    "ORTA":   PatternFill("solid", fgColor="FEF3C7"),
    "DÜŞÜK":  PatternFill("solid", fgColor="DBEAFE"),
}
GRADE_FILL = {
    "E": PatternFill("solid", fgColor="D9B3B3"), "D": PatternFill("solid", fgColor="F8CBCB"),
    "C": PatternFill("solid", fgColor="FCE0C8"), "B": PatternFill("solid", fgColor="FEF3C7"),
    "A": PatternFill("solid", fgColor="D1FAE5"),
}


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel(results: list, years: tuple, deep: bool,
                news: list = None) -> bytes:
    wb = openpyxl.Workbook()
    now = datetime.now().strftime("%d.%m.%Y %H:%M")

    # ---- 1) Yönetici Özeti ----
    ws = wb.active
    ws.title = "Yönetici Özeti"
    ws["A1"] = "KAP RİSK İZLEME RAPORU"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"Rapor tarihi: {now}  |  Kapsam: {min(years)}–{max(years)} "
                f"+ güncel dönem  |  Analiz derinliği: "
                f"{'Derin (bildirim tam metinleri)' if deep else 'Hızlı (başlık/özet)'}")
    ws["A3"] = ("Kaynak: KAP (kap.org.tr) kamuya açık bildirimleri. Skor 0-100; "
                "bulgu ağırlığı × güncellik katsayısı toplamıdır.")
    headers = ["Hisse", "Şirket Unvanı", "Risk Notu", "Risk Skoru",
               "Seviye", "Kritik", "Yüksek", "Orta", "Düşük",
               "Toplam Bulgu", "İyileşme Sinyali", "Taranan Bildirim",
               "Öne Çıkan Risk"]
    ws.append([])
    ws.append(headers)
    hrow = ws.max_row
    _style_header(ws, hrow, len(headers))
    for r in sorted(results, key=lambda x: -x["skor"]):
        counts = {s: sum(1 for f in r["findings"] if f["siddet"] == s)
                  for s in SEVERITY_ORDER}
        top = max(r["findings"], key=lambda f: f["agirlik"] * f["guncellik"],
                  default=None)
        ws.append(_xlrow([
            r["member"]["hisse"], r["member"]["unvan"], r["not"], r["skor"],
            r["seviye"], counts["KRİTİK"], counts["YÜKSEK"], counts["ORTA"],
            counts["DÜŞÜK"], len(r["findings"]), len(r["improvements"]),
            r["taranan"],
            f"{top['kategori']}: {top['baslik'][:60]}" if top else "—"]))
        row = ws.max_row
        ws.cell(row=row, column=3).fill = GRADE_FILL.get(r["not"],
                                                         GRADE_FILL["A"])
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = BORDER
    ws.freeze_panes = f"A{hrow + 1}"
    ws.auto_filter.ref = f"A{hrow}:{get_column_letter(len(headers))}{ws.max_row}"
    _autofit(ws, [9, 46, 9, 10, 9, 8, 8, 8, 8, 12, 14, 15, 60])

    # ---- 2) Bulgular ----
    ws2 = wb.create_sheet("Risk Bulguları")
    headers2 = ["Hisse", "Şirket", "Tarih", "Şiddet", "Kategori", "Ağırlık",
                "Güncellik", "Bildirim Başlığı", "Özet", "Tespit Gerekçesi",
                "Piyasa Geneli", "KAP Linki"]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))
    all_f = sorted(
        (f for r in results for f in r["findings"]),
        key=lambda f: (SEVERITY_ORDER.get(f["siddet"], 9), -f["tarih"].timestamp()
                       if f["tarih"] != datetime.min else 0))
    for f in all_f:
        ws2.append(_xlrow([f["hisse"], f["sirket"], f["tarih_str"], f["siddet"],
                           f["kategori"], f["agirlik"], f["guncellik"],
                           f["baslik"][:120], f["ozet"][:250], f["gerekce"],
                           "Evet" if f["piyasa_geneli"] else "Hayır", f["link"]]))
        row = ws2.max_row
        fill = SEV_FILL.get(f["siddet"])
        if fill:
            ws2.cell(row=row, column=4).fill = fill
        for c in range(1, len(headers2) + 1):
            ws2.cell(row=row, column=c).border = BORDER
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{ws2.max_row}"
    _autofit(ws2, [8, 34, 17, 9, 30, 8, 9, 45, 55, 40, 12, 40])

    # ---- 3) İyileşme Sinyalleri ----
    ws3 = wb.create_sheet("İyileşme Sinyalleri")
    ws3.append(["Hisse", "Tarih", "Kategori", "Başlık", "Gerekçe", "Link"])
    _style_header(ws3, 1, 6)
    for r in results:
        for f in r["improvements"]:
            ws3.append(_xlrow([f["hisse"], f["tarih_str"], f["kategori"],
                               f["baslik"][:120], f["gerekce"], f["link"]]))
    _autofit(ws3, [8, 17, 30, 50, 45, 40])

    # ---- 4) Kategori Analizi ----
    ws4 = wb.create_sheet("Kategori Analizi")
    if len(results) <= 30:     # az şirket: kategoriler satır, şirketler sütun
        ws4.append(["Risk Kategorisi", "Ağırlık"] +
                   [r["member"]["hisse"] for r in results] + ["TOPLAM"])
        _style_header(ws4, 1, 2 + len(results) + 1)
        for cat_id, (label, weight) in RISK_CATEGORIES.items():
            row = [label, weight]
            total = 0
            for r in results:
                n = sum(1 for f in r["findings"] if f["kategori_id"] == cat_id)
                row.append(n or "")
                total += n
            row.append(total)
            ws4.append(row)
        _autofit(ws4, [36, 9] + [8] * len(results) + [9])
    else:                      # çok şirket: şirketler satır, kategoriler sütun
        cat_ids = list(RISK_CATEGORIES)
        ws4.append(["Hisse", "Şirket"] +
                   [RISK_CATEGORIES[c][0] for c in cat_ids] + ["TOPLAM"])
        _style_header(ws4, 1, 2 + len(cat_ids) + 1)
        for r in sorted(results, key=lambda x: -x["skor"]):
            counts = {}
            for f in r["findings"]:
                counts[f["kategori_id"]] = counts.get(f["kategori_id"], 0) + 1
            ws4.append(_xlrow([r["member"]["hisse"], r["member"]["unvan"][:40]]
                              + [counts.get(c, "") for c in cat_ids]
                              + [len(r["findings"])]))
        ws4.freeze_panes = "A2"
        ws4.auto_filter.ref = (f"A1:{get_column_letter(2 + len(cat_ids) + 1)}"
                               f"{ws4.max_row}")
        _autofit(ws4, [8, 34] + [14] * len(cat_ids) + [9])

    # ---- 4b) Medya Sinyalleri ----
    if news:
        wsn = wb.create_sheet("Medya Sinyalleri")
        wsn.append(["Hisse", "Tarih", "Kaynak", "Haber Başlığı", "Link"])
        _style_header(wsn, 1, 5)
        for n in news:
            wsn.append(_xlrow([n["hisse"], n["tarih_str"],
                               n.get("kaynak") or "", n["baslik"][:200],
                               n["link"]]))
        wsn.freeze_panes = "A2"
        wsn.auto_filter.ref = f"A1:E{wsn.max_row}"
        _autofit(wsn, [8, 12, 18, 80, 45])

    # ---- 5) Metodoloji ----
    ws5 = wb.create_sheet("Metodoloji")
    lines = [
        ("KAP Risk İzleme — Metodoloji", True),
        ("", False),
        ("Veri kaynağı: KAP (kap.org.tr) kamuya açık şirket bildirimleri. "
         "Her şirket için seçilen yıllar + güncel dönem bildirim listesi "
         "çekilir; derin modda risk adayı bildirimlerin tam metinleri de "
         "indirilip taranır.", False),
        ("", False),
        ("Sınıflandırma: 15 risk kategorisinde Türkçe anahtar ifade "
         "sözlükleri ile kural tabanlı tarama yapılır. Derecelendirme "
         "bildirimlerinde not yönü (düşürme / teyit / yükseltme) tam metin "
         "üzerinden ayrıştırılır; yalnızca kötüleşme riskli sayılır. "
         "Kaldırma/çıkarma/lehte sonuç içeren bildirimler 'iyileşme "
         "sinyali' olarak ayrı raporlanır ve skoru etkilemez.", False),
        ("", False),
        ("Skor: her bulgu için kategori ağırlığı (1-10) × güncellik "
         "katsayısı (≤90 gün: 1.00, ≤1 yıl: 0.75, ≤2 yıl: 0.50, daha eski: "
         f"0.30) × {SCORE_GAIN} toplanır, 100 ile sınırlanır.", False),
        ("", False),
        ("Risk notları: A (0) Temiz · B (1-19) Düşük · C (20-44) Orta · "
         "D (45-69) Yüksek · E (70+) Kritik.", False),
        ("", False),
        ("Sınırlamalar: (1) PDF/Excel ekleri taranmaz; (2) kural tabanlı "
         "tespit nadir kalıpları kaçırabilir; (3) BIST/SPK piyasa geneli "
         "duyuruları ilgili tüm şirketlerde görünebilir ve ayrıca işaretlenir; "
         "(4) rapor yatırım tavsiyesi değildir.", False),
    ]
    for text, bold in lines:
        ws5.append([text])
        if bold:
            ws5.cell(row=ws5.max_row, column=1).font = TITLE_FONT
        ws5.cell(row=ws5.max_row, column=1).alignment = Alignment(wrap_text=True)
    ws5.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ═══════════════════════════════════════════════════════════ arayüz ════════


def inject_css():
    st.markdown("""
    <style>
      :root {
        --navy-900: #0b1220;
        --navy-800: #111827;
        --navy-700: #1f2a44;
        --navy-600: #243b63;
        --gray-950: #0f172a;
        --gray-700: #334155;
        --gray-500: #64748b;
        --gray-300: #cbd5e1;
        --gray-200: #e2e8f0;
        --gray-100: #f1f5f9;
        --gray-050: #f8fafc;
        --white: #ffffff;
      }

      html, body, [class*="css"] {
        font-family: Inter, ui-sans-serif, system-ui, -apple-system,
          BlinkMacSystemFont, "Segoe UI", sans-serif;
        color: var(--gray-950);
      }

      [data-testid="stAppViewContainer"] {
        background: var(--gray-050);
      }
      [data-testid="stHeader"] {background: transparent;}
      [data-testid="stMain"] .block-container {
        padding-top: 1.25rem;
        padding-bottom: 2.5rem;
        max-width: 1440px;
      }

      h1, h2, h3 {
        letter-spacing: 0;
        color: var(--gray-950);
      }
      h3 {
        font-size: 1.05rem;
        font-weight: 700;
        margin-top: 1.35rem;
      }
      p, li, .stMarkdown {
        line-height: 1.55;
      }

      [data-testid="stSidebar"] {
        background: var(--navy-900);
        border-right: 1px solid rgba(226,232,240,.12);
      }
      [data-testid="stSidebar"] [data-testid="stSidebarContent"] {
        padding-top: 1rem;
      }
      [data-testid="stSidebar"] * {color: #e5e7eb;}
      [data-testid="stSidebar"] h1,
      [data-testid="stSidebar"] h2,
      [data-testid="stSidebar"] h3 {
        color: #f8fafc;
        font-size: .98rem;
        margin-bottom: .45rem;
      }
      [data-testid="stSidebar"] .stCaption, [data-testid="stSidebar"] small
        {color: #a7b3c7 !important;}
      [data-testid="stSidebar"] label,
      [data-testid="stSidebar"] p {
        font-size: .86rem;
      }
      [data-testid="stSidebar"] [data-baseweb="tag"]
        {background: var(--navy-600); border-radius: 6px;}
      [data-testid="stSidebar"] [data-baseweb="select"] > div,
      [data-testid="stSidebar"] [data-baseweb="input"] > div,
      [data-testid="stSidebar"] input
        {background: #111827 !important; color: #f8fafc !important;
         border-color: #2f3b52 !important; border-radius: 7px !important;}
      [data-testid="stSidebar"] .stButton > button {
        min-height: 2.4rem;
      }

      .hero {
        background: var(--white);
        border: 1px solid var(--gray-200);
        border-left: 5px solid var(--navy-600);
        padding: 18px 24px;
        border-radius: 8px;
        color: var(--gray-950);
        margin-bottom: 18px;
        box-shadow: 0 8px 24px rgba(15,23,42,.05);
      }
      .hero-eyebrow {
        color: var(--navy-600);
        font-size: .76rem;
        font-weight: 800;
        letter-spacing: 0;
        margin-bottom: 5px;
        text-transform: uppercase;
      }
      .hero h1 {
        margin: 0;
        font-size: clamp(1.35rem, 2vw, 1.9rem);
        line-height: 1.2;
        letter-spacing: 0;
        color: var(--gray-950);
      }
      .hero p {
        margin: 7px 0 0 0;
        color: var(--gray-500);
        font-size: .92rem;
        max-width: 980px;
      }

      .system-status-grid {
        display: grid;
        grid-template-columns: repeat(4, minmax(0, 1fr));
        gap: 10px;
        margin: -6px 0 18px 0;
      }
      .system-status-card {
        background: var(--white);
        border: 1px solid var(--gray-200);
        border-radius: 8px;
        padding: 13px 15px;
        box-shadow: 0 4px 14px rgba(15,23,42,.04);
      }
      .system-status-label {
        color: var(--gray-500);
        font-size: .72rem;
        font-weight: 780;
        line-height: 1.25;
        text-transform: uppercase;
      }
      .system-status-value {
        color: var(--gray-950);
        font-size: 1.02rem;
        font-weight: 760;
        margin-top: 5px;
      }
      .system-status-dot {
        display: inline-block;
        width: 8px;
        height: 8px;
        border-radius: 999px;
        margin-right: 7px;
        background: #16a34a;
        vertical-align: 1px;
      }

      div[data-testid="stMetric"] {
        background: var(--white);
        border: 1px solid var(--gray-200);
        border-left: 3px solid var(--navy-600);
        padding: 13px 15px;
        border-radius: 8px;
        box-shadow: 0 4px 14px rgba(15,23,42,.04);
      }
      div[data-testid="stMetric"] label {
        color: var(--gray-500) !important;
        font-size: .76rem !important;
        font-weight: 700 !important;
        text-transform: uppercase;
      }
      div[data-testid="stMetricValue"] {
        color: var(--gray-950);
        font-weight: 760;
      }

      .stTabs [data-baseweb="tab-list"] {gap: 6px;}
      .stTabs [data-baseweb="tab"] {
        font-weight: 700;
        background: var(--white);
        border-radius: 7px;
        padding: 7px 14px;
        border: 1px solid var(--gray-200);
      }
      .stTabs [aria-selected="true"] {
        background: var(--navy-700) !important;
        color: #ffffff !important;
        border-color: var(--navy-700) !important;
      }
      .stTabs [data-baseweb="tab-highlight"], .stTabs [data-baseweb="tab-border"]
        {display: none;}

      [data-testid="stMain"] details[data-testid="stExpander"] {
        background: var(--white);
        border: 1px solid var(--gray-200) !important;
        border-radius: 8px !important;
        overflow: hidden;
        box-shadow: 0 2px 10px rgba(15,23,42,.03);
      }

      [data-testid="stSidebar"] [data-testid="stAlert"],
      [data-testid="stSidebar"] [data-testid="stAlert"] *
        {color: #7f1d1d !important;}

      [data-testid="stDataFrame"] {
        border-radius: 8px;
        overflow: hidden;
        border: 1px solid var(--gray-200);
      }
      .stButton > button[kind="primary"], .stDownloadButton > button {
        background: var(--navy-700);
        border: 1px solid var(--navy-700);
        border-radius: 7px;
        font-weight: 750;
      }
      .stButton > button[kind="primary"]:hover,
      .stDownloadButton > button:hover {
        background: var(--navy-600);
        border-color: var(--navy-600);
      }
      .empty-card {
        background: var(--white);
        border: 1px dashed var(--gray-300);
        border-radius: 8px;
        padding: 28px;
        text-align: center;
        color: var(--gray-700);
        margin-top: 8px;
      }
      .empty-card b {color: var(--navy-600);}

      .risk-card, .news-card {
        background: var(--white);
        border: 1px solid var(--gray-200);
        border-radius: 8px;
        padding: 14px 16px;
        margin-bottom: 10px;
        box-shadow: 0 2px 10px rgba(15,23,42,.035);
      }
      .risk-card {
        border-left: 4px solid var(--gray-500);
      }
      .news-card {
        border-left: 4px solid var(--navy-600);
      }
      .card-topline {
        display: flex;
        justify-content: space-between;
        gap: 12px;
        flex-wrap: wrap;
        align-items: baseline;
      }
      .card-company {
        font-weight: 760;
        color: var(--navy-700);
      }
      .card-date {
        color: var(--gray-500);
        font-size: .84rem;
      }
      .badge-row {
        display: flex;
        gap: 6px;
        flex-wrap: wrap;
        margin: 7px 0 5px 0;
      }
      .badge {
        border-radius: 6px;
        font-size: .72rem;
        font-weight: 760;
        padding: 2px 8px;
        line-height: 1.45;
      }
      .badge-muted {
        background: var(--gray-100);
        color: var(--gray-700);
      }
      .badge-new {
        background: #e8f5ee;
        color: #166534;
      }
      .badge-source {
        color: var(--gray-500);
        font-size: .76rem;
        align-self: center;
      }
      .card-title {
        color: var(--gray-950);
        font-weight: 720;
        margin-top: 7px;
      }
      .card-summary {
        color: var(--gray-700);
        font-size: .9rem;
        margin-top: 3px;
      }
      .card-link {
        display: inline-block;
        color: var(--navy-600);
        font-size: .84rem;
        font-weight: 760;
        margin-top: 7px;
        text-decoration: none;
      }
      .card-link:hover {
        text-decoration: underline;
      }

      @media (max-width: 720px) {
        [data-testid="stMain"] .block-container {
          padding-left: 1rem;
          padding-right: 1rem;
        }
        .hero {padding: 15px 16px;}
        .system-status-grid {
          grid-template-columns: 1fr;
          margin-top: -4px;
        }
        .stTabs [data-baseweb="tab"] {
          padding: 6px 10px;
          font-size: .84rem;
        }
      }
    </style>""", unsafe_allow_html=True)


def hero():
    st.markdown(f"""
    <div class="hero">
      <div class="hero-eyebrow">Kurumsal Risk İzleme</div>
      <h1>KAP Erken Uyarı</h1>
      <p>Temerrüt, yakın izleme, yeniden yapılandırma, regülatör cezası ve
         derecelendirme sinyalleri · {_format_tr_time(datetime.now(timezone.utc))}
         · kap.org.tr canlı verisi</p>
    </div>""", unsafe_allow_html=True)


def _to_istanbul(value) -> datetime | None:
    """UTC/aware/naive değeri Istanbul saatine çevir; hatada None döndür."""
    try:
        if isinstance(value, str):
            value = datetime.fromisoformat(value)
        if isinstance(value, datetime):
            if value.tzinfo is None:
                value = value.replace(tzinfo=timezone.utc)
            return value.astimezone(ISTANBUL_TZ)
    except Exception:
        return None
    return None


def _format_tr_time(value, fallback: str = "Henüz yok") -> str:
    """UI için Türkiye formatında tarih/saat göster."""
    dt = _to_istanbul(value)
    if dt:
        return dt.strftime("%d.%m.%Y %H:%M")
    return fallback


def _format_tr_clock(value, fallback: str = "30 dk içinde") -> str:
    dt = _to_istanbul(value)
    if dt:
        return dt.strftime("%H:%M")
    return fallback


def _latest_scan_from_file() -> datetime | None:
    """Opsiyonel latest_scan.json varsa son tarama zamanını UI'da kullan."""
    try:
        path = os.path.join(os.path.dirname(__file__), "latest_scan.json")
        with open(path, encoding="utf-8") as fh:
            data = json.load(fh)
        raw = data.get("ts") or data.get("timestamp") or data.get("last_scan")
        if raw:
            return datetime.fromisoformat(str(raw))
    except Exception:
        return None
    return None


def _system_status_values() -> dict:
    """Üst durum kartları için değerleri session state'ten güvenli şekilde oku."""
    try:
        meta = st.session_state.get("scan_meta") or {}
        last_scan = meta.get("ts") or _latest_scan_from_file()
        next_scan = None
        if isinstance(last_scan, datetime):
            next_scan = last_scan + timedelta(seconds=AUTO_REFRESH_SECONDS)
        elif isinstance(last_scan, str):
            next_scan = datetime.fromisoformat(last_scan) + timedelta(
                seconds=AUTO_REFRESH_SECONDS)
        return {
            "system": "Aktif",
            "last_scan": _format_tr_time(last_scan),
            "next_scan": (_format_tr_time(next_scan)
                          if next_scan else "30 dk içinde"),
            "actions": "Çalışıyor",
        }
    except Exception:
        return {
            "system": "Aktif",
            "last_scan": "Henüz yok",
            "next_scan": "30 dk içinde",
            "actions": "Çalışıyor",
        }


def render_system_status(container=None):
    vals = _system_status_values()
    html = f"""
    <div class="system-status-grid">
      <div class="system-status-card">
        <div class="system-status-label">Sistem Durumu</div>
        <div class="system-status-value">
          <span class="system-status-dot"></span>{html_lib.escape(vals["system"])}
        </div>
      </div>
      <div class="system-status-card">
        <div class="system-status-label">Son Tarama</div>
        <div class="system-status-value">{html_lib.escape(vals["last_scan"])}</div>
      </div>
      <div class="system-status-card">
        <div class="system-status-label">Sonraki Tarama</div>
        <div class="system-status-value">{html_lib.escape(vals["next_scan"])}</div>
      </div>
      <div class="system-status-card">
        <div class="system-status-label">GitHub Actions</div>
        <div class="system-status-value">{html_lib.escape(vals["actions"])}</div>
      </div>
    </div>"""
    target = container if container is not None else st
    target.markdown(html, unsafe_allow_html=True)


AUTO_FIRST_SCAN_SECONDS = 60      # açılıştan sonra ilk otomatik derin tarama
AUTO_REFRESH_SECONDS = 3600       # sonrasında saat başı yenileme


@st.fragment(run_every="20s")
def _auto_scan_ticker():
    """Açılışın 1. dakikasında otomatik derin tarama başlatır, ardından
    her saat başı taramayı yeniler. Geri sayımı kenar çubuğunda gösterir."""
    now = time.time()
    boot = st.session_state.setdefault("boot_ts", now)
    if not st.session_state.get("auto_ready", True):
        # kullanıcı elle seçim yapıyor ve liste boş: tetikleme, bekle
        st.caption("Otomatik tarama için şirket seçin")
        return
    if st.session_state.get("auto_pending"):
        st.caption("Otomatik tarama sırada...")
        return
    meta = st.session_state.get("scan_meta")
    if meta is None:
        remain = AUTO_FIRST_SCAN_SECONDS - (now - boot)
        if remain <= 0:
            st.session_state["auto_pending"] = True
            st.rerun(scope="app")
        else:
            st.caption(f"Otomatik derin tarama **{int(remain)} sn** "
                       "içinde başlayacak")
    else:
        elapsed = (datetime.now() - meta["ts"]).total_seconds()
        if elapsed >= AUTO_REFRESH_SECONDS:
            st.session_state["auto_pending"] = True
            st.rerun(scope="app")
        else:
            nxt = (meta["ts"] + timedelta(seconds=AUTO_REFRESH_SECONDS))
            st.caption(f"Sonraki otomatik yenileme: **{_format_tr_clock(nxt)}**")


def executive_summary_text(results):
    no_data = [r for r in results if r["seviye"] in (
        "VERİ ALINAMADI", "VERİ EKSİK / KONTROL GEREKLİ", "HATA")]
    results = [r for r in results if r not in no_data]
    total_f = sum(len(r["findings"]) for r in results)
    crit = [r for r in results if r["not"] in ("D", "E")]
    watch = [r for r in results if r["not"] == "C"]
    clean = [r for r in results if r["not"] == "A"]
    lines = [
        f"Taranan **{len(results)}** şirkette, **"
        f"{sum(r['taranan'] for r in results):,}** KAP bildirimi analiz edildi "
        f"ve **{total_f}** risk sinyali tespit edildi."]
    if crit:
        top = sorted(crit, key=lambda r: -r["skor"])
        names = ", ".join(f"**{r['member']['hisse']}** ({r['skor']:.0f}/100)"
                          for r in top)
        lines.append(f"🔴 **Acil dikkat gerektiren şirketler:** {names}. "
                     "Bu şirketlerde temerrüt, yakın izleme, yeniden "
                     "yapılandırma veya eşdeğer kritik sinyaller mevcut.")
        for r in top[:3]:
            worst = max(r["findings"],
                        key=lambda f: f["agirlik"] * f["guncellik"],
                        default=None)
            if worst:
                lines.append(f"&nbsp;&nbsp;• **{r['member']['hisse']}** — "
                             f"{worst['kategori']}: *{worst['baslik'][:80]}* "
                             f"({worst['tarih_str'][:10]})")
    if watch:
        lines.append("🟠 **İzleme listesi:** " +
                     ", ".join(r["member"]["hisse"] for r in watch) +
                     " — orta seviyeli sinyaller; periyodik takip önerilir.")
    if clean:
        lines.append("🟢 **Temiz:** " +
                      ", ".join(r["member"]["hisse"] for r in clean) +
                      " — taranan dönemde risk sinyali bulunmadı.")
    yeni = sum(1 for r in results for f in r["findings"] if f.get("yeni"))
    if yeni:
        lines.append(f"🆕 Bir önceki taramada bulunmayan **{yeni}** yeni "
                     "risk sinyali tespit edildi — Bildirim Akışı "
                     "sekmesinde 'Sadece yeni' filtresiyle görüntülenebilir.")
    impr = sum(len(r["improvements"]) for r in results)
    if impr:
        lines.append(f"↗️ Ayrıca **{impr}** iyileşme sinyali (tedbir kaldırma, "
                     "pazar çıkışı, lehte sonuç vb.) tespit edildi; detaylar "
                     "raporun ilgili bölümündedir.")
    if no_data:
        lines.append("⚠️ **Veri Eksik / Kontrol Gerekli:** " +
                     ", ".join(r["member"]["hisse"] for r in no_data) +
                     " — KAP erişimi geçici olarak başarısız oldu; bu "
                     "şirketler temiz kabul edilmedi, önceki başarılı veri "
                     "varsa korunur.")
    return "\n\n".join(lines)


def _load_state() -> dict:
    """Bulgu anahtarı → ilk görülme zamanı (ISO) sözlüğü."""
    try:
        with open(STATE_FILE, encoding="utf-8") as fh:
            data = json.load(fh)
        raw = data.get("keys", {})
        if isinstance(raw, list):          # eski liste biçiminden geçiş
            ts = data.get("ts", datetime.now().isoformat())
            return {k: ts for k in raw}
        return raw
    except Exception:
        return {}


def _load_prev() -> tuple:
    """(önceki anahtar kümesi, önceki tarama zamanı | None)."""
    keys = set(_load_state())
    ts = None
    try:
        with open(STATE_FILE, encoding="utf-8") as fh:
            raw_ts = json.load(fh).get("ts")
        if raw_ts:
            ts = datetime.fromisoformat(raw_ts)
    except Exception:
        pass
    return keys, ts


def _save_keys(keys: set):
    """Anahtarları ilk görülme zamanıyla ATOMİK yaz; 1 yıldan eskileri buda.

    Geçici dosya + os.replace: yazım yarıda kesilirse eski dosya bozulmaz,
    geçmiş kaybolmaz.
    """
    try:
        now = datetime.now()
        state = _load_state()
        for k in keys:
            state.setdefault(k, now.isoformat())
        cutoff = (now - timedelta(days=365)).isoformat()
        state = {k: t for k, t in state.items() if t >= cutoff}
        tmp = STATE_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as fh:
            json.dump({"keys": state, "ts": now.isoformat()}, fh)
        os.replace(tmp, STATE_FILE)
    except Exception:
        pass                       # salt-okunur dosya sistemi vb. — kritik değil


def _render_company_body(r):
    """Bir şirketin detay panelini (metrikler + bulgu tablosu) çizer."""
    d1, d2, d3, d4 = st.columns(4)
    d1.metric("Risk Skoru", f"{r['skor']:.0f} / 100")
    d2.metric("Risk Notu", f"{r['not']} ({r['seviye']})")
    d3.metric("Taranan Bildirim", r["taranan"])
    d4.metric("İyileşme Sinyali", len(r["improvements"]))
    if str(r["member"].get("islem", "")) == "0":
        st.caption("⛔ Bu üyenin payları borsada işlem görmüyor "
                   "(kodsuz ihraççı veya işlemi durdurulmuş şirket).")
    if r["findings"]:
        fdf = pd.DataFrame([{
            "Tarih": f["tarih_str"][:10],
            "Şiddet": f"{SEVERITY_EMOJI.get(f['siddet'], '')} {f['siddet']}",
            "Kategori": f["kategori"],
            "Başlık": f["baslik"][:70],
            "Özet": f["ozet"][:90],
            "Gerekçe": f["gerekce"],
            "Link": f["link"],
        } for f in sorted(r["findings"],
                          key=lambda f: (SEVERITY_ORDER.get(f["siddet"], 9),
                                         -f["tarih"].timestamp()
                                         if f["tarih"] != datetime.min else 0))])
        st.dataframe(fdf, use_container_width=True, hide_index=True,
                     column_config={"Link": st.column_config.LinkColumn(
                         "KAP", display_text="Aç")})
    elif r.get("veri_hatasi") and r["taranan"] == 0:
        st.warning("KAP erişimi geçici olarak başarısız oldu; şirket temiz "
                   "kabul edilmedi, önceki başarılı veri varsa korunur.")
    else:
        st.success("Taranan dönemde risk sinyali tespit edilmedi.")
    if r["improvements"]:
        st.markdown("**↗️ İyileşme sinyalleri:**")
        for f in r["improvements"]:
            st.markdown(f"- {f['tarih_str'][:10]} · {f['kategori']} · "
                        f"{f['baslik'][:70]} — {f['gerekce']} "
                        f"[KAP]({f['link']})")


_SEV_CARD = {"KRİTİK": ("#7f1d1d", "#fee2e2"), "YÜKSEK": ("#dc2626", "#ffedd5"),
             "ORTA": ("#d97706", "#fef3c7"), "DÜŞÜK": ("#2563eb", "#dbeafe")}


def _feed_card(f) -> str:
    """Tek bulgu için KAP görünümlü HTML kart (tıklayınca KAP'ta açılır)."""
    fg, bg = _SEV_CARD.get(f["siddet"], ("#334155", "#f1f5f9"))
    esc = html_lib.escape
    yeni = '<span class="badge badge-new">YENİ</span>' if f.get("yeni") else ""
    pg = ('<span class="badge-source">BIST/SPK piyasa duyurusu</span>'
          if f.get("piyasa_geneli") else "")
    ozet = esc(f["ozet"][:220]) if f["ozet"] else ""
    return f"""
<div class="risk-card" style="border-left-color:{fg};">
  <div class="card-topline">
    <span class="card-company">{esc(f["hisse"])} —
      {esc(f["sirket"][:55])}</span>
    <span class="card-date">{esc(f["tarih_str"])}</span>
  </div>
  <div class="badge-row">
    <span class="badge" style="background:{bg};color:{fg};">
      {esc(f["siddet"])}</span>
    <span class="badge badge-muted">{esc(f["kategori"])}</span>{yeni}{pg}
  </div>
  <div class="card-title">{esc(f["baslik"][:120])}</div>
  <div class="card-summary">{ozet}</div>
  <a class="card-link" href="{esc(f["link"])}" target="_blank">
     Bildirimi KAP'ta aç</a>
</div>"""


def _news_card(n) -> str:
    esc = html_lib.escape
    return f"""
<div class="news-card">
  <div class="card-topline">
    <span class="card-company">{esc(n["hisse"])}
      <span class="badge-source">· {esc(n.get("kaynak") or "medya")}</span></span>
    <span class="card-date">{esc(n["tarih_str"])}</span>
  </div>
  <div class="card-title">{esc(n["baslik"][:140])}</div>
  <a class="card-link" href="{esc(n["link"])}" target="_blank">Haberi aç</a>
</div>"""


# ═══════════════════════════════════════════════ talep üzerine e-posta ════

EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
MAIL_BEKLEME_SN = 60          # oturum başına gönderimler arası bekleme
MAIL_TEST_BEKLEME_SN = 30     # eksiz test gönderimleri arası bekleme
MAIL_GUNLUK_TAVAN = 100       # tüm ziyaretçiler için günlük gönderim tavanı
MAIL_SAYAC_DOSYA = "rapor_mail_sayaci.json"

MAIL_SECRET_ALIASES = {
    "SMTP_HOST": ("SMTP_HOST", "SMTP_SERVER", "MAIL_HOST"),
    "SMTP_PORT": ("SMTP_PORT", "MAIL_PORT"),
    "SMTP_USER": ("SMTP_USER", "SMTP_USERNAME", "GMAIL_USER",
                  "EMAIL_USER", "MAIL_USER"),
    "SMTP_PASS": ("SMTP_PASS", "SMTP_PASSWORD", "GMAIL_APP_PASSWORD",
                  "GMAIL_APP_PASS", "GMAIL_PASSWORD", "EMAIL_PASSWORD",
                  "MAIL_PASSWORD"),
    "SMTP_OAUTH_TOKEN": ("SMTP_OAUTH_TOKEN", "OAUTH_TOKEN",
                         "GMAIL_OAUTH_TOKEN", "GOOGLE_OAUTH_TOKEN"),
    "MAIL_FROM": ("MAIL_FROM", "SMTP_FROM", "EMAIL_FROM", "FROM_EMAIL",
                  "MAIL_SENDER"),
}

MAIL_SECRET_SECTIONS = ("smtp", "SMTP", "mail", "MAIL", "email", "EMAIL",
                        "gmail", "GMAIL")


def _secret_deger(secrets, canonical: str):
    aliases = MAIL_SECRET_ALIASES[canonical]
    top_level_keys = [key for alias in aliases
                      for key in (alias, alias.lower())]
    for key in top_level_keys:
        try:
            value = secrets.get(key)
        except Exception:
            value = secrets[key] if key in secrets else None
        if value:
            return value

    nested_keys = {k.lower() for k in aliases}
    nested_keys |= {canonical.lower(), canonical.lower().replace("smtp_", "")}
    if canonical == "SMTP_PASS":
        nested_keys |= {"pass", "password", "app_password", "app_pass"}
    elif canonical == "SMTP_USER":
        nested_keys |= {"user", "username", "login"}
    elif canonical == "SMTP_HOST":
        nested_keys |= {"host", "server"}
    elif canonical == "SMTP_PORT":
        nested_keys |= {"port"}
    elif canonical == "MAIL_FROM":
        nested_keys |= {"from", "sender"}

    for section_name in MAIL_SECRET_SECTIONS:
        try:
            section = secrets.get(section_name)
        except Exception:
            section = secrets[section_name] if section_name in secrets else None
        if not section:
            continue
        for key in nested_keys:
            try:
                value = section.get(key)
            except Exception:
                try:
                    value = section[key]
                except Exception:
                    value = None
            if value:
                return value
    return None


def _mail_secrets_env():
    """Streamlit Cloud'da SMTP sırları st.secrets'tan env'e köprülenir."""
    try:
        for k in MAIL_SECRET_ALIASES:
            if not os.environ.get(k):
                value = _secret_deger(st.secrets, k)
                if value:
                    os.environ[k] = str(value)
    except Exception:
        pass


def _mail_tavan_oku() -> dict:
    bugun = date.today().isoformat()
    try:
        with open(MAIL_SAYAC_DOSYA, encoding="utf-8") as fh:
            veri = json.load(fh)
    except Exception:
        veri = {}
    if veri.get("tarih") != bugun:
        veri = {"tarih": bugun, "sayi": 0}
    return veri


def _mail_tavan_uygun() -> bool:
    return _mail_tavan_oku().get("sayi", 0) < MAIL_GUNLUK_TAVAN


def _mail_tavan_bump():
    veri = _mail_tavan_oku()
    veri["sayi"] = veri.get("sayi", 0) + 1
    try:
        with open(MAIL_SAYAC_DOSYA, "w", encoding="utf-8") as fh:
            json.dump(veri, fh)
    except Exception:
        pass


def _mail_gonderen(mailer) -> str:
    host = mailer._env("SMTP_HOST", mailer.VARSAYILAN_SMTP_HOST).lower()
    user = mailer._env("SMTP_USER")
    mail_from = mailer._env("MAIL_FROM")
    # Gmail SMTP'de From adresi giriş yapılan hesapla aynı olmazsa bazı
    # kurumsal alıcılar mesajı sessizce karantinaya alabiliyor.
    if "gmail" in host and user:
        return user
    return mail_from or user or mailer.VARSAYILAN_FROM


def talep_mail_gonder(alici: str, results, years, deep, xls,
                      ekleri_ekle: bool = True) -> tuple:
    """Mevcut tarama raporunu girilen adrese gönderir."""
    alici = (alici or "").strip()
    if not EMAIL_RE.match(alici):
        return False, "Geçerli bir e-posta adresi girin."
    if not results:
        return False, "Önce bir tarama çalıştırın; gönderilecek rapor yok."
    son = st.session_state.get("_mail_son_ts", 0.0)
    kalan = MAIL_BEKLEME_SN - (time.time() - son)
    if kalan > 0:
        return False, f"Çok sık talep. {int(kalan)} sn sonra tekrar deneyin."
    if not _mail_tavan_uygun():
        return False, ("Günlük e-posta gönderim tavanına ulaşıldı; "
                       "raporu yukarıdaki düğmeden indirebilirsiniz.")

    import kap_rapor_mail as mailer
    _mail_secrets_env()
    try:
        excel = xls or build_excel(results, years, deep) if ekleri_ekle else b""
        satirlar = (
            [{k: f.get(k, "") for k in mailer.CSV_ALANLARI}
             for r in results for f in r["findings"]]
            if ekleri_ekle else []
        )
        csv_veri = mailer.csv_bytes(satirlar) if ekleri_ekle else b""
        konu, duz, html = mailer.mail_govdesi(results, years, deep)
        if not ekleri_ekle:
            konu += " - eksiz ozet"
            duz += ("\n\nNot: Bu mesaj kurumsal e-posta filtrelerine takılmaması "
                    "için ek dosya içermeyen özet olarak gönderildi.")
            html += ("<p style='color:#64748b;font-size:12px;"
                     "font-family:Segoe UI,Arial,sans-serif'>Bu mesaj kurumsal "
                     "e-posta filtrelerine takılmaması için ek dosya içermeyen "
                     "özet olarak gönderildi.</p>")
        gonderen = _mail_gonderen(mailer)
        stamp = datetime.now(ISTANBUL_TZ).strftime("%Y%m%d_%H%M")
        msg = mailer.mail_olustur(konu, duz, html, gonderen, [alici],
                                  excel, csv_veri, stamp,
                                  ekleri_ekle=ekleri_ekle)
        sonuc = mailer.gonder(msg)
    except SystemExit as exc:
        return False, (f"E-posta gönderilemedi (yapılandırma): {exc} "
                       "Yöneticiye SMTP_PASS ayarını hatırlatın.")
    except Exception as exc:
        return False, f"E-posta gönderilemedi: {exc}"

    st.session_state["_mail_son_ts"] = time.time()
    _mail_tavan_bump()
    msg_id = sonuc.get("message_id") if isinstance(sonuc, dict) else ""
    ek = f" Mesaj kimliği: {msg_id}" if msg_id else ""
    tur = "eksiz özet raporu" if not ekleri_ekle else "ekli raporu"
    return True, (f"SMTP sunucusu {tur} {alici} adresi için kabul etti."
                  f"{ek} Gmail dışı kutuda görünmüyorsa spam/karantina ve "
                  "kurumsal dış gönderici filtresini kontrol edin.")


def talep_test_mail_gonder(alici: str) -> tuple:
    """Ek dosyasız SMTP test maili gönderir."""
    alici = (alici or "").strip()
    if not EMAIL_RE.match(alici):
        return False, "Geçerli bir e-posta adresi girin."
    son = st.session_state.get("_mail_test_son_ts", 0.0)
    kalan = MAIL_TEST_BEKLEME_SN - (time.time() - son)
    if kalan > 0:
        return False, f"Çok sık test. {int(kalan)} sn sonra tekrar deneyin."

    import kap_rapor_mail as mailer
    _mail_secrets_env()
    try:
        gonderen = _mail_gonderen(mailer)
        msg = mailer.test_mail_olustur(gonderen, [alici])
        sonuc = mailer.gonder(msg)
    except SystemExit as exc:
        return False, (f"Test e-postası gönderilemedi (yapılandırma): {exc} "
                       "SMTP_PASS ayarını kontrol edin.")
    except Exception as exc:
        return False, f"Test e-postası gönderilemedi: {exc}"

    st.session_state["_mail_test_son_ts"] = time.time()
    msg_id = sonuc.get("message_id") if isinstance(sonuc, dict) else ""
    ek = f" Mesaj kimliği: {msg_id}" if msg_id else ""
    return True, (f"Eksiz test maili {alici} adresi için SMTP tarafından "
                  f"kabul edildi.{ek}")



def render_dashboard(results, years, deep, news=None, date_range=None):
    ok = [r for r in results if "hata" not in r]
    no_data = [r for r in ok
               if r.get("veri_hatasi") and r["taranan"] == 0]
    partial = [r for r in ok
               if r.get("veri_hatasi") and r["taranan"] > 0]
    if no_data:
        st.warning(
            "**Veri Eksik / Kontrol Gerekli:** " +
            "; ".join(f"{r['member']['hisse']} için KAP erişimi geçici "
                      "olarak başarısız oldu" for r in no_data) +
            ". Şirket temiz kabul edilmedi; önceki başarılı veri varsa "
            "korunur. Birkaç dakika sonra taramayı yeniden çalıştırın.")
    if partial:
        st.warning(
            "**Veri Eksik / Kontrol Gerekli:** " +
            ", ".join(r["member"]["hisse"] for r in partial) +
            " için bazı sorgular/bildirim detayları alınamadı (KAP geçici "
            "kısıtlaması). Sonuçlar eksik olabilir; şirket temiz kabul "
            "edilmez. Birkaç dakika sonra yeniden tarama önerilir.")
    failed = [r for r in results if "hata" in r]
    if failed:
        st.warning("**Veri Eksik / Kontrol Gerekli:** " + ", ".join(
            f"{r['member']['hisse']} ({str(r.get('hata', ''))[:60]})"
            for r in failed) + " — taramayı yeniden çalıştırın.")
    if not ok:
        st.error("Hiçbir şirket taranamadı. Ağ bağlantınızı kontrol edip "
                 "**Taramayı Başlat**'a yeniden tıklayın.")
        return
    df_score = pd.DataFrame([{
        "Hisse": r["member"]["hisse"],
        "Şirket": r["member"]["unvan"][:45],
        "Not": r["not"], "Skor": r["skor"], "Seviye": r["seviye"],
        "Durum": f"{r['emoji']} {r['seviye']}",
        "Bulgu": len(r["findings"]),
        "İyileşme": len(r["improvements"]),
        "Taranan": r["taranan"],
    } for r in ok]).sort_values("Skor", ascending=False)

    tab1, tab_feed, tab2, tab3, tab4 = st.tabs(
        ["Yönetici Özeti", "Bildirim Akışı", "Şirket Detayları",
         "Tüm Bulgular", "Rapor İndir"])

    # ─── Bildirim akışı (KAP formatında kötü haber akışı) ───
    with tab_feed:
        feed = sorted((f for r in ok for f in r["findings"]),
                      key=lambda f: f["tarih"], reverse=True)
        fc1, fc2, fc3 = st.columns([2, 1, 1])
        f_sev = fc1.multiselect("Şiddet", list(SEVERITY_ORDER),
                                default=list(SEVERITY_ORDER), key="feed_sev")
        f_new = fc2.checkbox("Sadece yeni bulgular", key="feed_new",
                             help="Bir önceki taramada olmayan bulgular")
        f_lim = fc3.select_slider("Gösterilecek kayıt", [40, 80, 150, 300],
                                  value=80, key="feed_lim")
        shown = [f for f in feed if f["siddet"] in f_sev
                 and (not f_new or f.get("yeni"))]
        rng = (f" · {date_range[0].strftime('%d.%m.%Y')} – "
               f"{date_range[1].strftime('%d.%m.%Y')}") if date_range else ""
        st.caption(f"{len(shown)} riskli bildirim{rng} — karta tıklayıp "
                   "**KAP'ta aç** bağlantısıyla bildirimin aslına gidin.")
        if not shown:
            st.info("Seçili filtrelerle gösterilecek bildirim yok.")
        for f in shown[:f_lim]:
            st.markdown(_feed_card(f), unsafe_allow_html=True)
        if len(shown) > f_lim:
            st.caption(f"… ve {len(shown) - f_lim} kayıt daha "
                       "(tamamı Excel raporunda).")
        if news:
            st.markdown("---")
            st.markdown("### Medya Sinyalleri")
            st.caption("Ulusal basında risk temalı haberler — skora dahil "
                       "edilmez, teyit için KAP bildirimleriyle birlikte "
                       "değerlendirin.")
            for n in news[:60]:
                st.markdown(_news_card(n), unsafe_allow_html=True)

    # ─── Yönetici özeti ───
    with tab1:
        c1, c2, c3, c4, c5 = st.columns(5)
        allf = [f for r in ok for f in r["findings"]]
        c1.metric("Taranan Şirket", len(ok))
        c2.metric("Analiz Edilen Bildirim", f"{sum(r['taranan'] for r in ok):,}")
        c3.metric("Risk Sinyali", len(allf))
        c4.metric("Kritik/Yüksek Bulgu",
                  sum(1 for f in allf if f["siddet"] in ("KRİTİK", "YÜKSEK")))
        c5.metric("Riskli Şirket (D/E)",
                  sum(1 for r in ok if r["not"] in ("D", "E")))

        st.markdown("### Yönetici Özeti")
        st.markdown(executive_summary_text(ok))

        st.markdown("### Risk Skoru Sıralaması")
        chart_df = df_score.head(60)
        if len(df_score) > 60:
            st.caption(f"Grafikte en yüksek skorlu 60 şirket gösteriliyor "
                       f"(toplam {len(df_score)}); tamamı aşağıdaki skor "
                       "tablosunda ve Excel raporunda.")
        chart = alt.Chart(chart_df).mark_bar(cornerRadius=4).encode(
            x=alt.X("Skor:Q", scale=alt.Scale(domain=[0, 100]),
                    title="Risk Skoru (0-100)"),
            y=alt.Y("Hisse:N", sort="-x", title=None),
            color=alt.Color("Seviye:N", title="Seviye", scale=alt.Scale(
                domain=["KRİTİK", "YÜKSEK", "ORTA", "DÜŞÜK", "TEMİZ",
                        "VERİ EKSİK / KONTROL GEREKLİ"],
                range=["#7f1d1d", "#dc2626", "#ea580c", "#ca8a04",
                       "#16a34a", "#64748b"])),
            tooltip=["Hisse", "Şirket", "Skor", "Not", "Bulgu"],
        ).properties(height=max(220, 34 * len(chart_df)))
        st.altair_chart(chart, use_container_width=True)

        colA, colB = st.columns(2)
        with colA:
            st.markdown("### Kategori Isı Haritası")
            heat_set = set(chart_df["Hisse"])
            heat_rows = [{"Hisse": r["member"]["hisse"],
                          "Kategori": f["kategori"], "Adet": 1}
                         for r in ok for f in r["findings"]
                         if r["member"]["hisse"] in heat_set]
            if heat_rows:
                hdf = (pd.DataFrame(heat_rows)
                       .groupby(["Hisse", "Kategori"]).sum().reset_index())
                heat = alt.Chart(hdf).mark_rect().encode(
                    x=alt.X("Hisse:N", title=None),
                    y=alt.Y("Kategori:N", title=None),
                    color=alt.Color("Adet:Q", scale=alt.Scale(scheme="reds")),
                    tooltip=["Hisse", "Kategori", "Adet"],
                ).properties(height=340)
                st.altair_chart(heat, use_container_width=True)
            else:
                st.info("Bulgu yok.")
        with colB:
            st.markdown("### Zaman Çizelgesi")
            tl_rows = [{"Tarih": f["tarih"], "Hisse": f["hisse"],
                        "Şiddet": f["siddet"], "Başlık": f["baslik"][:60],
                        "Kategori": f["kategori"]}
                       for r in ok for f in r["findings"]
                       if f["tarih"] != datetime.min]
            tl_rows.sort(key=lambda x: x["Tarih"], reverse=True)
            if len(tl_rows) > 4000:      # tarayıcıyı boğmamak için son 4000
                st.caption(f"En güncel 4.000 bulgu gösteriliyor "
                           f"(toplam {len(tl_rows)}).")
                tl_rows = tl_rows[:4000]
            if tl_rows:
                tdf = pd.DataFrame(tl_rows)
                tl = alt.Chart(tdf).mark_circle(size=110).encode(
                    x=alt.X("Tarih:T", title=None),
                    y=alt.Y("Hisse:N", title=None),
                    color=alt.Color("Şiddet:N", scale=alt.Scale(
                        domain=["KRİTİK", "YÜKSEK", "ORTA", "DÜŞÜK"],
                        range=["#7f1d1d", "#dc2626", "#ea580c", "#2563eb"])),
                    tooltip=["Hisse", "Şiddet", "Kategori", "Başlık",
                             alt.Tooltip("Tarih:T", format="%d.%m.%Y")],
                ).properties(height=340)
                st.altair_chart(tl, use_container_width=True)
            else:
                st.info("Bulgu yok.")

        st.markdown("### Skor Tablosu")
        st.dataframe(df_score, use_container_width=True, hide_index=True,
                     column_config={"Skor": st.column_config.ProgressColumn(
                         "Skor", min_value=0, max_value=100, format="%.0f")})

    # ─── Şirket detayları ───
    with tab2:
        ranked = sorted(ok, key=lambda x: -x["skor"])
        top50, rest = ranked[:50], ranked[50:]
        if rest:
            st.caption(f"Skor sırasıyla ilk 50 şirket listeleniyor; kalan "
                       f"{len(rest)} şirketi alttaki kutudan seçebilirsiniz.")
        for r in top50:
            m = r["member"]
            with st.expander(
                    f"{r['emoji']} **{m['hisse']}** — {m['unvan'][:60]}  ·  "
                    f"Not: **{r['not']}** · Skor: **{r['skor']:.0f}/100** · "
                    f"{len(r['findings'])} bulgu",
                    expanded=(r["not"] in ("D", "E") and len(top50) <= 15)):
                _render_company_body(r)
        if rest:
            st.markdown("---")
            pick = st.selectbox(
                "Diğer şirketler",
                [f"{r['member']['hisse']} — {r['member']['unvan'][:50]} "
                 f"(skor {r['skor']:.0f})" for r in rest],
                index=None, placeholder="Şirket seçin...")
            if pick:
                code = pick.split(" — ")[0]
                r = next(x for x in rest if x["member"]["hisse"] == code)
                _render_company_body(r)

    # ─── Tüm bulgular ───
    with tab3:
        allf = [f for r in ok for f in r["findings"]]
        if allf:
            fc1, fc2, fc3 = st.columns(3)
            sev_sel = fc1.multiselect("Şiddet", list(SEVERITY_ORDER),
                                      default=list(SEVERITY_ORDER))
            cat_sel = fc2.multiselect(
                "Kategori", sorted({f["kategori"] for f in allf}))
            co_sel = fc3.multiselect(
                "Şirket", sorted({f["hisse"] for f in allf}))
            rows = [f for f in allf
                    if f["siddet"] in sev_sel
                    and (not cat_sel or f["kategori"] in cat_sel)
                    and (not co_sel or f["hisse"] in co_sel)]
            bdf = pd.DataFrame([{
                "Yeni": "Evet" if f.get("yeni") else "",
                "Hisse": f["hisse"], "Tarih": f["tarih_str"],
                "Şiddet": f["siddet"], "Kategori": f["kategori"],
                "Başlık": f["baslik"][:80], "Özet": f["ozet"][:120],
                "Gerekçe": f["gerekce"],
                "Piyasa Geneli": "Evet" if f["piyasa_geneli"] else "",
                "Link": f["link"],
            } for f in sorted(rows, key=lambda f: (SEVERITY_ORDER.get(f["siddet"], 9),
                                                   -f["tarih"].timestamp()
                                                   if f["tarih"] != datetime.min else 0))])
            st.dataframe(bdf, use_container_width=True, hide_index=True, height=560,
                         column_config={"Link": st.column_config.LinkColumn(
                             "KAP", display_text="Aç")})
        else:
            st.info("Seçilen şirketlerde risk bulgusu yok.")

    # ─── İndir ───
    with tab4:
        st.markdown("### Yapılandırılmış Risk Yönetimi Raporu")
        st.markdown(
            "Excel raporu 5 sayfadan oluşur: **Yönetici Özeti** (skor kartı), "
            "**Risk Bulguları** (şiddet renklendirmeli, filtrelenebilir), "
            "**İyileşme Sinyalleri**, **Kategori Analizi** (şirket × kategori "
            "matrisi) ve **Metodoloji**.")
        xls = st.session_state.get("xlsx")
        if xls is None:               # tarama sonrası üretim başarısızsa
            try:
                xls = build_excel(ok, years, deep, news=news)
                st.session_state["xlsx"] = xls
            except Exception as exc:
                st.error(f"Excel raporu üretilemedi: {exc}")
                xls = None
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        if xls:
            st.download_button(
                "Excel Raporu İndir (.xlsx)", data=xls,
                file_name=f"KAP_Risk_Raporu_{stamp}.xlsx",
                mime=("application/vnd.openxmlformats-officedocument"
                      ".spreadsheetml.sheet"),
                type="primary")
        allf = [f for r in ok for f in r["findings"]]
        if allf:
            csv_df = pd.DataFrame([{k: (v if not isinstance(v, datetime) else
                                        v.strftime("%d.%m.%Y %H:%M"))
                                    for k, v in f.items() if k != "tarih"}
                                   for f in allf])
            st.download_button(
                "Ham Bulgular (.csv)",
                data=csv_df.to_csv(index=False).encode("utf-8-sig"),
                file_name=f"KAP_Risk_Bulgular_{stamp}.csv", mime="text/csv")

        st.divider()
        st.markdown("### Raporu E-posta ile İste")
        st.caption(
            "Mevcut tarama raporunu Excel ve CSV ekleriyle doğrudan "
            "e-posta adresinize gönderebilirsiniz. Kurumsal e-posta "
            "adreslerinde önce eksiz test mailini deneyin.")
        with st.form("rapor_mail_talep", clear_on_submit=False):
            alici = st.text_input("E-posta adresi",
                                  placeholder="umutokan.basbay@halkbank.com.tr")
            c_mail, c_ozet, c_test = st.columns([1, 1, 1])
            gonder_btn = c_mail.form_submit_button("Ekli Raporu Gönder")
            ozet_btn = c_ozet.form_submit_button("Eksiz Özet Gönder")
            test_btn = c_test.form_submit_button("Test Maili Gönder")
        if test_btn:
            with st.spinner("Eksiz test maili gönderiliyor..."):
                basarili, mesaj = talep_test_mail_gonder(alici)
            (st.success if basarili else st.warning)(mesaj)
        if ozet_btn:
            with st.spinner("Eksiz özet raporu gönderiliyor..."):
                basarili, mesaj = talep_mail_gonder(alici, ok, years,
                                                     deep, xls,
                                                     ekleri_ekle=False)
            (st.success if basarili else st.warning)(mesaj)
        if gonder_btn:
            with st.spinner("Rapor gönderiliyor..."):
                basarili, mesaj = talep_mail_gonder(alici, ok, years,
                                                     deep, xls,
                                                     ekleri_ekle=True)
            (st.success if basarili else st.warning)(mesaj)


def main():
    st.set_page_config(page_title="KAP Risk İzleme Platformu",
                       page_icon="K", layout="wide")
    inject_css()
    hero()
    system_status_slot = st.empty()
    render_system_status(system_status_slot)

    # ── kenar çubuğu: parametreler ──
    with st.sidebar:
        st.header("Tarama Parametreleri")
        try:
            # rehberi oturum boyunca sabitle: 24 saatlik önbellek tazelenince
            # seçenek etiketleri kayıp kullanıcı seçimini sıfırlamasın
            if "dir_df" not in st.session_state:
                st.session_state["dir_df"] = fetch_member_directory()
            directory = st.session_state["dir_df"]
        except Exception as exc:
            st.error(f"KAP şirket listesi alınamadı: {exc}")
            # otomatik döngü kilitlenmesin: bayrağı temizle, zamanlayıcıyı
            # kayıtlı tut — bir sonraki tikte yeniden denenir
            st.session_state.pop("auto_pending", None)
            _auto_scan_ticker()
            st.stop()
        options, opt_to_oid, oid_to_opt = [], {}, {}
        for row in directory.itertuples():
            kod = row.kodlar if len(str(row.kodlar)) <= 18 else row.hisse
            label = f"{kod} — {row.unvan[:48]}"
            options.append(label)
            opt_to_oid[label] = row.oid
            oid_to_opt[row.oid] = label

        default_oids, media_terms = resolve_default_members(directory)
        default_opts = [oid_to_opt[o] for o in default_oids
                        if o in oid_to_opt]

        preset_on = st.checkbox(
            f"Varsayılan izleme listesi ({len(default_opts)} üye)",
            value=True,
            help="Tiki kaldırırsanız listeden dilediğiniz üyeleri "
                 "ekleyip çıkarabilirsiniz.")
        select_all = st.checkbox("Tümünü seç "
                                 f"({len(options)} KAP üyesi)")
        if preset_on or select_all:
            manual = default_opts
        else:
            manual = st.multiselect(
                "Şirketler", options, default=default_opts,
                key="company_sel",
                placeholder="Hisse kodu veya unvan yazın...")
        selected_opts = (options if select_all
                         else (default_opts if preset_on else manual))
        with st.expander("İzleme evreni"):
            st.caption(f"KAP rehberi: **{len(directory)}** üye (kodsuz "
                       "ihraççılar `*` işaretli). KAP üyesi olmayan "
                       f"**{len(media_terms)}** grup yalnızca medya "
                       "taramasında izlenir: " + ", ".join(media_terms))

        # tarihi oturum boyunca sabitle: gece yarısı geçişinde widget
        # kimliği değişip kullanıcının aralığı sıfırlanmasın
        today = st.session_state.setdefault("session_today", date.today())
        dr = st.date_input(
            "Tarama aralığı (elle tarih girebilirsiniz)",
            value=(date(today.year - 2, 1, 1), today),
            min_value=date(EARLIEST_YEAR, 1, 1), max_value=today,
            format="DD.MM.YYYY",
            help="KAP tarih filtresini sunucu tarafında desteklemediği "
                 "için ilgili yıllar çekilir, aralık yerelde uygulanır.")
        if isinstance(dr, tuple) and len(dr) == 2:
            d_start, d_end = dr
        elif isinstance(dr, tuple) and len(dr) == 1:
            d_start, d_end = dr[0], today          # ikinci tarih seçilmedi
        elif isinstance(dr, tuple):                # () — alan temizlendi
            d_start, d_end = date(today.year - 2, 1, 1), today
        else:                                      # tek date nesnesi
            d_start, d_end = dr, today
        years = tuple(range(d_start.year, d_end.year + 1))

        deep = st.radio(
            "Analiz derinliği",
            ["Derin — bildirim tam metinleri de taranır (önerilen)",
             "Hızlı — yalnızca başlık ve özet"],
            index=0).startswith("Derin")

        news_on = st.checkbox(
            "İnternet haberlerini de tara (Google News)", value=True,
            help="Her şirket için ulusal basında temerrüt/konkordato/"
                 "yasak vb. temalı haberler aranır. Skora dahil edilmez, "
                 "bilgilendirme amaçlıdır.")

        run = st.button("Taramayı Başlat", type="primary",
                        use_container_width=True,
                        disabled=not selected_opts)
        st.session_state["auto_ready"] = bool(selected_opts)
        _auto_scan_ticker()

    auto_due = st.session_state.pop("auto_pending", False)
    if (run or auto_due) and selected_opts:
        sel_oids = {opt_to_oid[o] for o in selected_opts if o in opt_to_oid}
        members = directory[directory.oid.isin(sel_oids)].to_dict("records")
        results = []
        # bitiş "bugün" seçiliyse gece yarısı sonrası taramalarda gerçek
        # bugüne genişlet (widget kimliği için tarih donduruldu)
        eff_end = date.today() if d_end >= today else d_end
        date_range = (d_start, eff_end)
        prog = st.progress(0.0)
        status = st.status(f"{len(members)} şirket taranıyor...",
                           expanded=True)
        for i, m in enumerate(members, 1):
            status.write(f"[{i}/{len(members)}] **{m['hisse']}** — "
                         f"{m['unvan'][:50]}")
            try:
                results.append(scan_company(m, years, deep, date_range))
            except Exception as exc:
                status.write(f"&nbsp;&nbsp;⚠️ {m['hisse']} taranamadı: {exc}")
                results.append({"member": m, "taranan": 0, "findings": [],
                                "improvements": [], "skor": 0.0, "not": "-",
                                "seviye": "VERİ EKSİK / KONTROL GEREKLİ",
                                "emoji": "⚠️",
                                "renk": "#666", "hata": str(exc)})
            prog.progress(i / len(members))

        # 🆕 yeni bulgu işaretleme: anahtar kararlı oid ile kurulur ve
        # yalnızca YAYIN TARİHİ son taramadan yeni bildirimler işaretlenir.
        # Böylece izlemeye yeni eklenen bir şirketin yıllar önceki
        # bildirimleri "yeni" diye parlamaz; anahtar biçimi değişse bile
        # eski kayıtlar toplu 🆕 yağmuruna dönmez.
        prev_keys, prev_ts = _load_prev()
        horizon = (prev_ts - timedelta(days=3)) if prev_ts else None
        cur_keys = set()
        for r in results:
            for f in r["findings"]:
                key = f"{f.get('oid') or f['hisse']}:{f['bildirim_no']}"
                cur_keys.add(key)
                f["yeni"] = (bool(prev_keys) and key not in prev_keys
                             and f["tarih"] != datetime.min
                             and (horizon is None or f["tarih"] >= horizon))
        _save_keys(prev_keys | cur_keys)

        # 📡 medya taraması (skora girmez; en riskli 40 şirketle sınırlı,
        # paralel çekim — Google, KAP hız sınırlayıcısından bağımsız)
        news = []
        if news_on:
            targets = sorted((r for r in results if "hata" not in r),
                             key=lambda x: -x["skor"])[:40]
            status.write(f"{len(targets)} şirket + "
                         f"{len(media_terms)} KAP-dışı grup için medya "
                         "taraması...")
            jobs = [(company_short_name(r["member"]["unvan"]),
                     r["member"]["hisse"], r["member"]["unvan"])
                    for r in targets]
            jobs += [(term, term, f"{term} (KAP üyesi değil — yalnız medya)")
                     for term in media_terms]
            with ThreadPoolExecutor(max_workers=8) as ex:
                futs = {ex.submit(fetch_company_news, q): (h, s)
                        for q, h, s in jobs}
                for fut in as_completed(futs):
                    h, s = futs[fut]
                    try:
                        items = fut.result()
                    except Exception:
                        continue
                    for n in items:
                        if n["dt"] != datetime.min and not (
                                date_range[0] <= n["dt"].date()
                                <= date_range[1]):
                            continue
                        news.append({**n, "hisse": h, "sirket": s})
            news.sort(key=lambda x: x["dt"], reverse=True)

        status.update(label="Tarama tamamlandı ✅", state="complete",
                      expanded=False)
        prog.empty()
        st.session_state["results"] = results
        st.session_state["news"] = news
        st.session_state["scan_meta"] = {"years": years, "deep": deep,
                                         "range": date_range,
                                         "ts": datetime.now()}
        # Excel'i tarama biter bitmez BİR KEZ üret; her widget
        # etkileşiminde yeniden kurulmasın
        try:
            st.session_state["xlsx"] = build_excel(
                [r for r in results if "hata" not in r], years, deep,
                news=news)
        except Exception:
            st.session_state["xlsx"] = None

    render_system_status(system_status_slot)

    if "results" in st.session_state:
        meta = st.session_state["scan_meta"]
        render_dashboard(st.session_state["results"], meta["years"],
                         meta["deep"], st.session_state.get("news", []),
                         meta.get("range"))
    else:
        st.markdown("""
        <div class="empty-card">
          <p style="margin:0 0 4px 0;font-size:1.05rem;">
            Varsayılan izleme listesi için <b>otomatik derin tarama</b>
            birazdan başlayacak.</p>
          <p style="margin:0;font-size:.9rem;color:#64748b;">
            Beklemek istemiyorsanız soldan <b>Taramayı Başlat</b>'a
            tıklayabilirsiniz.</p>
        </div>""", unsafe_allow_html=True)


if __name__ == "__main__":
    main()
